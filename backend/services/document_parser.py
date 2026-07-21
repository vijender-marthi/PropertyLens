"""Parse uploaded PDFs and Excel files to extract financial data."""
import json
import logging
import re
from pathlib import Path
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)


def _looks_spaceless(text: str) -> bool:
    """Heuristic: pdfminer/pdfplumber sometimes drop the spaces between words
    on certain statement PDFs, gluing words into 'PropertyAddress',
    'AccountNumber', 'OutstandingPrincipal'. The tell-tale is an internal
    lower→upper case transition, which correctly spaced English text almost
    never has — so we use that rather than word length (legalese has plenty of
    legitimately long words like 'electronically')."""
    if not text:
        return True
    camel = re.findall(r'[a-z][A-Z]', text)
    return len(camel) >= 8


def _strip_md_tables(md: str) -> str:
    """Flatten MarkItDown's markdown tables into plain 'label value' lines so
    the field parser can read them. Drops separator rows and pipe scaffolding."""
    out = []
    for line in md.splitlines():
        s = line.strip()
        if re.fullmatch(r'\|?[\s|:\-]+\|?', s) and '-' in s:
            continue  # table separator row (| --- | --- |)
        if '|' in s:
            cells = [c.strip() for c in s.split('|') if c.strip()]
            if cells:
                out.append(' '.join(cells))
        else:
            out.append(line)
    return '\n'.join(out)


def markitdown_convert(filepath: str) -> str:
    """Convert a document to readable markdown with Microsoft MarkItDown."""
    from markitdown import MarkItDown
    return MarkItDown().convert(filepath).text_content or ""


def _markitdown_text(filepath: str) -> str:
    try:
        return _strip_md_tables(markitdown_convert(filepath))
    except Exception as exc:
        logger.warning("MarkItDown conversion failed for %s: %s", filepath, exc)
        return ""


def _log_loan_document_extraction_gaps(category: str, raw_data: Dict[str, Any], text: str, filepath: str) -> None:
    if category == "1098":
        missing = [
            label for key, label in (
                ("mortgage_interest", "Box 1 interest"),
                ("current_balance", "Box 2 balance"),
                ("property_address", "property address"),
            )
            if raw_data.get(key) in (None, "")
        ]
    elif category == "mortgage_statement":
        missing = [
            label for key, label in (
                ("current_balance", "current balance"),
                ("property_address", "property address"),
            )
            if raw_data.get(key) in (None, "")
        ]
    else:
        missing = []
    if missing:
        logger.warning(
            "%s extraction missing %s from MarkItDown text for %s. Text preview: %s",
            category,
            ", ".join(missing),
            filepath,
            (text or "")[:2000],
        )


def _pdfplumber_text(filepath: str, x_tolerance: float = 3) -> str:
    import pdfplumber
    text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text(x_tolerance=x_tolerance, y_tolerance=3)
            if page_text:
                text += page_text + "\n"
    return text


def _ocr_text(filepath: str) -> str:
    """Render each PDF page at 300 DPI and OCR with Tesseract.

    Last resort for scanned / image-only PDFs where no text layer exists.
    Requires: PyMuPDF (fitz), pytesseract, Pillow.
    """
    try:
        import fitz                 # PyMuPDF
        import pytesseract
        from PIL import Image
        import io
        doc = fitz.open(filepath)
        pages_text = []
        for page in doc:
            mat = fitz.Matrix(300 / 72, 300 / 72)   # 300 DPI
            pix = page.get_pixmap(matrix=mat)
            img = Image.open(io.BytesIO(pix.tobytes('png')))
            pages_text.append(pytesseract.image_to_string(img))
        return '\n'.join(pages_text)
    except Exception:
        return ''


def _clean_ocr_artifacts(text: str) -> str:
    """Repair common OCR artifacts in scanned financial documents."""
    # OCR sometimes inserts spaces inside dollar amounts: "$1 65,984.70" → "$165,984.70"
    text = re.sub(r'(\$\s*\d{1,3})\s+(\d{2,3},\d{3}(?:\.\d{1,2})?)', lambda m: m.group(1).replace(' ', '') + m.group(2), text)
    # "3.625 %" or "3.625  %" → "3.625%"
    text = re.sub(r'(\d+\.?\d*)\s{1,3}%', r'\1%', text)
    return text


def _camel_count(text: str) -> int:
    return len(re.findall(r'[a-z][A-Z]', text)) if text else 10 ** 9


_ADJACENCY_VALUE_RE = re.compile(
    r'(\$\s*[\d,]+\.?\d*|\d+\.\d+\s*%|\d{1,2}/\d{1,2}/\d{2,4}|\d{2,3}\s+months)\s*$'
)
_STANDALONE_VALUE_RE = re.compile(
    r'(?m)^\s*(?:\$\s*[\d,]+\.\d{2}|\d+\.\d+\s*%|\d{1,2}/\d{1,2}/\d{2,4})\s*$'
)


def _label_value_adjacency(text: str) -> int:
    """Count lines where a text label is immediately followed, on the SAME
    line, by its value (money / date / percent / term).

    Two-column statements (label column + value column) are read out of order by
    pdfminer/MarkItDown — every value lands on its own line, separated from its
    label — so they score ~0 here. A correctly interleaved extraction
    ("Original Principal Balance $506,250.00") scores high. This lets us detect
    and reject a mis-ordered extraction that would otherwise break every
    label->value regex downstream.
    """
    if not text:
        return 0
    count = 0
    for line in text.splitlines():
        s = line.strip()
        m = _ADJACENCY_VALUE_RE.search(s)
        if m and re.match(r'[A-Za-z].*[A-Za-z]', s[:m.start()].strip()):
            count += 1
    return count


def extract_pdf_text(filepath: str) -> str:
    """Extract readable plain text from a PDF for field parsing.

    Priority order:
      1. MarkItDown (pdfminer-based) — best for text PDFs with embedded fonts
      2. pdfplumber — recovers spacing when pdfminer glues words
      3. Tesseract OCR — last resort for scanned / image-only PDFs

    The winning candidate is whichever has the fewest camelCase run-togethers.
    OCR output is passed through _clean_ocr_artifacts() before returning.
    """
    candidates = []

    md_text = _markitdown_text(filepath)
    if md_text and len(md_text.strip()) > 30:
        if not _looks_spaceless(md_text):
            # MarkItDown produced spaced text, but two-column statements (a
            # label column beside a value column) are read out of order: all
            # values first, then all labels. That breaks every label->value
            # regex. Detect the pathology (few label->value lines despite many
            # standalone values) and fall back to pdfplumber's positional
            # extraction, which reconstructs the correct row order.
            md_adjacency = _label_value_adjacency(md_text)
            standalone_values = len(_STANDALONE_VALUE_RE.findall(md_text))
            if not (md_adjacency < 4 and standalone_values >= 6):
                return md_text      # Fast path: well-ordered text from MarkItDown
            for xt in (3, 1):
                try:
                    t = _pdfplumber_text(filepath, x_tolerance=xt)
                    if t and len(t.strip()) > 30 and not _looks_spaceless(t):
                        candidates.append(t)
                except Exception:
                    pass
            # Prefer the extraction with the best label->value adjacency; break
            # ties on the fewest camelCase run-togethers.
            return max(
                [md_text, *candidates],
                key=lambda c: (_label_value_adjacency(c), -_camel_count(c)),
            )
        candidates.append(md_text)
    else:
        logger.warning("MarkItDown produced empty text for PDF upload %s; trying text/OCR fallbacks.", filepath)

    # pdfplumber fallback with two tolerances
    for xt in (3, 1):
        try:
            t = _pdfplumber_text(filepath, x_tolerance=xt)
            if t and len(t.strip()) > 30:
                candidates.append(t)
        except Exception as e:
            if not candidates:
                candidates.append(f"Error extracting PDF: {e}")

    non_empty = [c for c in candidates if c and len(c.strip()) > 30]
    if non_empty:
        return min(non_empty, key=_camel_count)

    # Nothing worked — scanned / image-only PDF: fall back to OCR
    ocr = _ocr_text(filepath)
    if ocr and len(ocr.strip()) > 30:
        return _clean_ocr_artifacts(ocr)
    logger.warning("Document text extraction produced no readable text for %s", filepath)
    return '\n'.join(c for c in candidates if c) or ''


def _norm_header(value) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())


SPREADSHEET_COLUMN_ALIASES = {
    "property": "property_name",
    "lender": "lender_name",
    "residencytype": "residency_type",
    "originalresidencystatus": "original_residency_status",
    "currentresidencystatus": "current_residency_status",
    "primatedaterange": "primary_date_range",
    "primarydaterange": "primary_date_range",
    "rentaldaterange": "rental_date_range",
    "recordeddate": "recorded_date",
    "heldperiod": "held_period",
    "hoaflag": "hoa_flag",
    "solarlease": "solar_ownership",
    "loanaccountnumber": "account_number",
    "loannumber": "account_number",
    "accountnumber": "account_number",
    "loanproduct": "loan_product",
    "product": "loan_product",
    "estimatedtotalmonthlypayment": "estimated_total_monthly_payment",
    "totpayment": "estimated_total_monthly_payment",
    "totalpayment": "estimated_total_monthly_payment",
    "originalltv": "original_ltv",
    "escrowincluded": "escrow_included",
    "loanoriginationmonth": "origination_date",
    "loanoriginationdate": "origination_date",
    "loanacquisitiondate": "origination_date",
    "settlementdateorpurchasedate": "purchase_date",
    "settlementdate": "purchase_date",
    "purchasedate": "purchase_date",
    "salepriceakapurchaseprice": "purchase_price",
    "saleprice": "purchase_price",
    "totalamount": "settlement_total_amount",
    "closingcosts": "closing_costs",
    "begmonth": "period_start",
    "beginningmonth": "period_start",
    "endingmonth": "period_end",
    "endmonth": "period_end",
    "months": "months",
    "calendaryear": "tax_year",
    "year": "tax_year",
    "mortgageinterestreceived": "mortgage_interest",
    "mortgageinterest": "mortgage_interest",
    "outstandingmortgageprincipal": "current_balance",
    "outstandingprincipal": "current_balance",
    "unpaidprincipalbalance": "current_balance",
    "mortgageinsurancepremiums": "mortgage_insurance",
    "pmi": "mortgage_insurance",
    "insurance": "annual_insurance",
    "pointspaidonpurchaseofprincipalresidence": "points_paid",
    "points": "points_paid",
    "interestpaidytd": "interest_paid_ytd",
    "principalpaidytd": "principal_paid_ytd",
    "projectedprincipalfy": "projected_principal_fy",
    "projectedinterestfy": "projected_interest_fy",
    "mortgagetenurecovered": "mortgage_tenure_covered",
    "pointspaidonpurchaseofprincipalresidence": "points_paid",
    "propertytaxespaid": "property_tax_amount",
    "realestatetaxespaid": "property_tax_amount",
    "homeinsurance": "home_insurance",
    "yearendoutstandingbalance": "year_end_outstanding_balance",
    "principalpaiddown": "principal_paid_down",
    "cumprincipalpaiddown": "cumulative_principal_paid",
    "cumulativeprincipalpaiddown": "cumulative_principal_paid",
    "expectedprincipalpaiddown": "expected_principal_paid",
    "expectedprincipalpaid": "expected_principal_paid",
    "topup": "principal_topup_paid",
    "topuppaid": "principal_topup_paid",
    "extraprincipalpaid": "principal_topup_paid",
}


def _coerce_spreadsheet_value(key: str, value):
    if value is None or value == "":
        return None
    if key in {
        "mortgage_interest", "current_balance", "mortgage_insurance",
        "points_paid", "property_tax_amount", "home_insurance",
        "year_end_outstanding_balance", "principal_paid_down",
        "cumulative_principal_paid", "expected_principal_paid",
        "principal_topup_paid",
    }:
        if isinstance(value, (int, float)):
            return round(float(value), 2)
        return parse_currency(str(value))
    if key in {"tax_year", "months"}:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return value
    return value


def _parse_spreadsheet_rows(rows: list) -> list:
    for idx, row in enumerate(rows):
        aliases = [SPREADSHEET_COLUMN_ALIASES.get(_norm_header(c)) for c in row]
        if sum(1 for a in aliases if a) >= 4:
            parsed = []
            for raw in rows[idx + 1:]:
                item = {}
                for col, alias in enumerate(aliases):
                    if not alias or col >= len(raw):
                        continue
                    val = _coerce_spreadsheet_value(alias, raw[col])
                    if val is not None:
                        item[alias] = val
                if item:
                    parsed.append(item)
            return parsed
    return []


def extract_excel_data(filepath: str) -> Dict[str, Any]:
    """Extract raw sheets plus normalized mortgage/tax rows from Excel."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        data = {}
        parsed_rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            data[sheet_name] = rows
            parsed_rows.extend(_parse_spreadsheet_rows(rows))
        if parsed_rows:
            data["parsed_rows"] = parsed_rows
            latest = parsed_rows[-1]
            for key in (
                "property_name", "lender_name", "residency_type",
                "account_number", "origination_date", "tax_year",
                "mortgage_interest", "current_balance",
                "property_tax_amount", "year_end_outstanding_balance",
                "principal_paid_down", "cumulative_principal_paid",
                "expected_principal_paid", "principal_topup_paid",
            ):
                if latest.get(key) is not None:
                    data[key] = latest[key]
        return data
    except Exception as e:
        return {"error": str(e)}


def detect_category(text: str) -> str:
    """Guess the document category from its text content."""
    t = text.lower()
    if re.search(r'\bannual\s+statement\b', t) and re.search(r'transaction\s+activity', t) and re.search(r'amount\s+disbursed\s+from\s+escrow|county\s+taxes|homeowners?\s+insurance\s+premium', t):
        return 'escrow_analysis'
    if re.search(r'annual\s+escrow\s+account|escrow\s+account\s+(?:analysis|disclosure)|closer\s+look\s+at\s+your\s+escrow|statement\s+of\s+activity\s+in\s+your\s+escrow\s+account', t):
        return 'escrow_analysis'
    # CFPB Loan Estimates routinely say "compare with your Closing
    # Disclosure." That reference must not reclassify the estimate as a final
    # Closing Disclosure. Prefer an exact document-title line.
    if re.search(r'^\s*loan\s+estimate\s*$', text, re.IGNORECASE | re.MULTILINE):
        return 'loan_disclosure'
    # All closing / settlement documents → closing_statement
    # This covers: CFPB Closing Disclosure, ALTA Settlement Statement, HUD-1
    if re.search(r'^\s*closing\s+disclosure\s*$', text, re.IGNORECASE | re.MULTILINE):
        # A refinance Closing Disclosure establishes a new debt and belongs in
        # Loan Setup. Purchase Closing Disclosures continue through Property
        # Setup because they also establish purchase price and closing costs.
        if re.search(r'\bpurpose\s+refinance\b', t):
            return 'loan_disclosure'
        return 'closing_statement'
    if re.search(r'alta\s+settlement\s+statement|hud-1\s+settlement\s+statement', t):
        return 'closing_statement'
    if re.search(r'sale\s+price\s+of\s+property.*settlement\s+date|settlement\s+date.*sale\s+price\s+of\s+property', t, re.DOTALL):
        return 'closing_statement'
    if re.search(r'loan\s+estimate|loan\s+disclosure', t):
        return 'loan_disclosure'
    # Servicer "Loan Information" / online mortgage statements routinely embed a
    # Form 1098 year-end snapshot. Classify them by the statement body (live
    # balance, rate and next payment) rather than the embedded 1098 box list,
    # which would otherwise drop the loan terms. A standalone 1098 has none of
    # these live-servicing fields.
    if (
        re.search(r'current\s+principal\s+balance|loan\s+information\s+as\s+of', t)
        and re.search(r'(?:current\s+annual\s+)?interest\s+rate', t)
        and re.search(r'(?:next|monthly|regular)\s+payment(?:\s+amount)?', t)
    ):
        return 'mortgage_statement'
    # 1098/1099 forms must be matched before tax_return: their boilerplate
    # references "Form 1040 / Schedule A", which otherwise trips tax_return.
    if re.search(r'form\s*1098|mortgage\s+interest\s+statement|mortgage\s+interest\s+received\s+from', t):
        return '1098'
    if re.search(r'form\s*1099|\b1099\b', t):
        return '1099'
    if re.search(r'schedule\s*e|form\s*1040|tax\s+return', t):
        return 'tax_return'
    if re.search(r'mortgage\s+statement|escrow\s+balance|principal\s+balance', t):
        return 'mortgage_statement'
    if re.search(r'bank\s+statement|account\s+summary|checking\s+account|savings\s+account|beginning\s+balance|ending\s+balance', t):
        return 'bank_statement'
    if re.search(r'property\s+tax\s+statement|property\s+taxes?\s+for|assessed\s+value|parcel\s+#|full\s+cash\s+value|limited\s+value', t):
        return 'property_tax'
    if re.search(r'(?:homeowners?|hazard)\s+insurance|insurance\s+(?:declarations?|policy)|declarations?\s+page', t):
        return 'insurance_declaration'
    return 'other'


def parse_escrow_analysis(text: str) -> Dict[str, Any]:
    """Parse an escrow analysis as an auditable ledger, not one annual value."""
    from datetime import datetime

    data: Dict[str, Any] = parse_property_address(text)

    def amount(label: str, source: str = text) -> Optional[float]:
        match = re.search(label + r'[^\n$]*\$\s*([\d,]+\.\d{2})', source, re.IGNORECASE)
        return float(match.group(1).replace(',', '')) if match else None

    def labeled_date(label: str) -> Optional[str]:
        match = re.search(label + r'\s*:?\s*([A-Za-z]+\s+\d{1,2},\s*\d{4}|\d{1,2}/\d{1,2}/\d{4})', text, re.IGNORECASE)
        return _normalize_date(match.group(1)) if match else None

    def month_date(value: str, end: bool = False) -> str:
        value = value.strip()
        for fmt in ('%B %Y', '%b %Y', '%m/%Y'):
            try:
                parsed = datetime.strptime(value, fmt)
                if end:
                    import calendar
                    return parsed.replace(day=calendar.monthrange(parsed.year, parsed.month)[1]).strftime('%Y-%m-%d')
                return parsed.replace(day=1).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return _normalize_date(value) or value

    def category(description: str) -> str:
        normalized = re.sub(r'[^A-Z0-9]+', ' ', description.upper()).strip()
        if 'COUNTY TAX' in normalized or 'PROPERTY TAX' in normalized:
            return 'PROPERTY_TAX'
        if 'HOMEOWNERS INS' in normalized or 'HAZ INS' in normalized or 'HAZARD INS' in normalized:
            return 'HOMEOWNERS_INSURANCE'
        if 'BEGINNING BALANCE' in normalized:
            return 'BEGINNING_BALANCE'
        if 'TRANSFER BAL' in normalized:
            return 'ESCROW_TRANSFER'
        if 'SHORTAGE' in normalized:
            return 'SHORTAGE_PAYMENT'
        if 'OVERAGE' in normalized or 'REFUND' in normalized:
            return 'OVERAGE_REFUND'
        if 'DEPOSIT' in normalized:
            return 'ESCROW_DEPOSIT'
        return 'OTHER'

    def parse_money(value: str) -> Optional[float]:
        if value is None:
            return None
        raw = str(value).strip()
        if not raw or raw in {'—', '-'}:
            return None
        negative = raw.startswith('(') and raw.endswith(')')
        raw = raw.strip('()').replace('$', '').replace(',', '').strip()
        try:
            parsed = float(raw)
        except ValueError:
            return None
        return -parsed if negative else parsed

    def annual_activity_statement() -> Optional[Dict[str, Any]]:
        if not (
            re.search(r'\bAnnual\s+Statement\b', text, re.IGNORECASE)
            and re.search(r'Transaction\s+Activity', text, re.IGNORECASE)
        ):
            return None
        annual: Dict[str, Any] = parse_property_address(text)
        loan = re.search(r'Loan\s+Number\s*:\s*([A-Z0-9X*\-]+)', text, re.IGNORECASE)
        if loan:
            annual['loan_number'] = annual['account_number'] = loan.group(1).strip()
        servicer = re.search(r'\bServicer\s*:\s*([^\n]+)', text, re.IGNORECASE)
        if servicer:
            annual['servicer'] = servicer.group(1).strip()
        elif re.search(r'PENNYMAC|Pennymac', text):
            annual['servicer'] = 'Pennymac'
        notice = re.search(r'Notice\s+Date\s*:\s*([A-Za-z]+\s+\d{1,2},\s*\d{4})', text, re.IGNORECASE)
        if notice:
            annual['statement_date'] = _normalize_date(notice.group(1))
        period = re.search(
            r'(?:from|activity\s*\(?)\s*(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),\s*((?:19|20)\d{2})\s*(?:to|-)\s*(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),\s*((?:19|20)\d{2})',
            text,
            re.IGNORECASE,
        )
        if period:
            start_text = f'{period.group(1)} {period.group(2)}, {period.group(3)}'
            end_text = f'{period.group(4)} {period.group(5)}, {period.group(6)}'
            annual['history_period_start'] = _normalize_date(start_text)
            annual['history_period_end'] = _normalize_date(end_text)
            annual['expense_year'] = int(period.group(3))
            annual['statement_year'] = int(period.group(3))
        address = re.search(r'Property\s+Address\s*:\s*([^\n]+)\n\s*([^\n]+)', text, re.IGNORECASE)
        if address:
            annual['property_address'] = re.split(
                r'\s+(?=Credits\s+since\s+last\s+statement|Debits\s+since\s+last\s+statement|Payment\s+Activity|Transaction\s+Activity)\b',
                address.group(1).strip(),
                flags=re.IGNORECASE,
            )[0].strip(' ,')
            locality = re.search(r'(.+?)[,\s]+([A-Z]{2})\s+(\d{5})(?:-\d{4})?', address.group(2).strip())
            if locality:
                annual['property_city'] = locality.group(1).strip(' ,')
                annual['property_state'] = locality.group(2)
                annual['property_zip'] = locality.group(3)
        summary_fields = (
            ('current_balance', r'Unpaid\s+Principal\s+Balance\s+as\s+of\s+12-31-\d{4}'),
            ('mortgage_interest', r'Amount\s+Paid\s+to\s+Interest'),
            ('principal_paid_ytd', r'Aggregate\s+of\s+payments\s+applied\s+towards\s+principal'),
            ('principal_interest_paid_ytd', r'Amount\s+Paid\s+to\s+Principal\s+&\s+Interest\s+\(P&I\)'),
            ('escrow_deposited', r'Amount\s+Deposited\s+into\s+Escrow'),
            ('actual_total_disbursement', r'Amount\s+Disbursed\s+from\s+Escrow'),
            ('current_escrow_balance', r'Current\s+Escrow\s+Balance'),
        )
        for key, pattern in summary_fields:
            match = re.search(pattern + r'[^\n$()]*(\(?\$?[\d,]+\.\d{2}\)?)', text, re.IGNORECASE)
            if match:
                value = parse_money(match.group(1))
                if key == 'actual_total_disbursement' and value is not None:
                    value = abs(value)
                annual[key] = value

        activities = []
        latest_escrow_deposit = None
        for match in re.finditer(
            r'^\s*(\d{2}-\d{2}-\d{4})\s+(.+?)\s+(\(?\$[\d,]+\.\d{2}\)?)\s+(\(?\$[\d,]+\.\d{2}\)?)\s+(\(?\$[\d,]+\.\d{2}\)?)\s+(\(?\$[\d,]+\.\d{2}\)?)\s+(\(?\$[\d,]+\.\d{2}\)?)\s+(\(?\$[\d,]+\.\d{2}\)?)\s*$',
            text,
            re.MULTILINE,
        ):
            activity_date = _normalize_date(match.group(1))
            description = match.group(2).strip()
            transaction_amount = parse_money(match.group(3))
            principal = parse_money(match.group(4))
            interest = parse_money(match.group(5))
            escrow = parse_money(match.group(6))
            activity_type = category(description)
            item = {
                'activity_date': activity_date,
                'activity_type': activity_type,
                'source_description': description,
                'phase': 'HISTORICAL',
                'value_status': 'ACTUAL',
            }
            if activity_type in {'PROPERTY_TAX', 'HOMEOWNERS_INSURANCE'}:
                item['actual_disbursement'] = abs(transaction_amount or 0)
            elif activity_type == 'OVERAGE_REFUND':
                item['actual_disbursement'] = abs(transaction_amount or 0)
            elif description.lower().startswith('payment'):
                item['activity_type'] = 'ESCROW_DEPOSIT'
                item['actual_deposit'] = escrow if escrow and escrow > 0 else None
                if escrow and escrow > 0:
                    latest_escrow_deposit = escrow
            elif activity_type == 'OTHER':
                item['actual_deposit'] = transaction_amount if transaction_amount and transaction_amount > 0 else None
                item['actual_disbursement'] = abs(transaction_amount) if transaction_amount and transaction_amount < 0 else None
            if principal and principal > 0:
                item['principal'] = principal
            if interest and interest > 0:
                item['interest'] = interest
            activities.append(item)
        if activities:
            annual['activities'] = activities
            tax_total = round(sum(row.get('actual_disbursement') or 0 for row in activities if row.get('activity_type') == 'PROPERTY_TAX'), 2)
            insurance_total = round(sum(row.get('actual_disbursement') or 0 for row in activities if row.get('activity_type') == 'HOMEOWNERS_INSURANCE'), 2)
            if tax_total:
                annual['actual_tax'] = tax_total
            if insurance_total:
                annual['actual_insurance'] = insurance_total
            if latest_escrow_deposit is not None:
                annual['current_escrow_payment'] = latest_escrow_deposit
        annual['period_type'] = 'yearly'
        return {key: value for key, value in annual.items() if value is not None}

    annual_data = annual_activity_statement()
    if annual_data:
        return annual_data

    loan_match = re.search(r'Loan\s+Number\s*:\s*([A-Z0-9X*\-]+)', text, re.IGNORECASE)
    if loan_match:
        data['loan_number'] = data['account_number'] = loan_match.group(1).strip()
    data['servicer'] = (
        'Rocket Mortgage' if re.search(r'Rocket\s+Mortgage', text, re.IGNORECASE)
        else 'LoanCare' if re.search(r'LoanCare|lakeviewloanservicing', text, re.IGNORECASE)
        else None
    )

    address_match = re.search(r'Property\s+Address\s*:\s*([^\n]+)\n\s*([^\n]+)', text, re.IGNORECASE)
    if address_match:
        data['property_address'] = address_match.group(1).strip()
        locality = re.search(r'([^\n,]+),\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?', address_match.group(2))
        if locality:
            data['property_city'], data['property_state'], data['property_zip'] = (part.strip() for part in locality.groups())

    data['statement_date'] = labeled_date(r'Statement\s+Date')
    data['effective_date'] = labeled_date(r'(?:New\s+Payment\s+)?Effective\s+Date')
    if not data.get('statement_date'):
        header_date = re.search(r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}\b', text, re.IGNORECASE)
        if header_date:
            data['statement_date'] = _normalize_date(header_date.group(0))
    data['projected_minimum_balance'] = amount(r'Projected\s+Minimum\s+Balance')
    data['required_minimum_balance'] = amount(r'Required\s+Minimum\s+Balance')
    cushion = re.search(r'(?:selected\s+minimum\s+allowed\s+balance|cushion)\s+is\s+\$\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
    data['escrow_cushion'] = float(cushion.group(1).replace(',', '')) if cushion else None
    data['shortage_amount'] = amount(r'Shortage\s+Amount')
    data['overage_amount'] = amount(r'Overage\s+Amount')
    refund = re.search(r'(?:refund|check)[^\n$]{0,80}\$\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
    data['refund_amount'] = float(refund.group(1).replace(',', '')) if refund else data.get('overage_amount')

    decision = re.search(
        r'Principal\s*&\s*Interest:\s*([^\n]+)\n\s*Escrow\s+Payment:\s*([^\n]+)(?:\n\s*Shortage:\s*([^\n]+))?\n\s*Monthly\s+Payment:\s*([^\n]+)',
        text, re.IGNORECASE,
    )
    if decision:
        rows = [[float(item.replace(',', '')) for item in re.findall(r'\$\s*([\d,]+\.\d{2})', group)] for group in decision.groups()]
        if rows[0]: data['principal_interest_payment'] = rows[0][0]
        if rows[1]:
            data['current_escrow_payment'] = rows[1][0]
            data['new_escrow_payment'] = rows[1][-1]
        if rows[3]:
            data['current_total_payment'] = rows[3][0]
            data['new_total_payment'] = rows[3][-1]
        if len(rows[1]) >= 3 and len(rows[2]) >= 2 and rows[2][-1] > 0:
            data['selected_payment_option'] = 'OPTION_B_WITH_SHORTAGE'
    if data.get('current_escrow_payment') is None:
        escrow_line = re.search(r'^\s*Escrow\s+Payment:\s*([^\n]+)', text, re.IGNORECASE | re.MULTILINE)
        escrow_values = [float(item.replace(',', '')) for item in re.findall(r'\$\s*([\d,]+\.\d{2})', escrow_line.group(1))] if escrow_line else []
        if escrow_values:
            data['current_escrow_payment'] = escrow_values[0]
            data['new_escrow_payment'] = escrow_values[-1]
    if data.get('current_escrow_payment') is None:
        data['current_escrow_payment'] = amount(r'Current\s+Escrow\s+Payment')
    if data.get('current_escrow_payment') is None:
        legacy_payment = re.search(r'monthly\s+mortgage\s+payment[^$]*\$\s*([\d,]+\.\d{2})[^$]*\$\s*([\d,]+\.\d{2})\s+was\s+for\s+principal\s+and\s+interest[^$]*\$\s*([\d,]+\.\d{2})\s+went\s+into\s+your\s+escrow', text, re.IGNORECASE | re.DOTALL)
        if legacy_payment:
            data['current_total_payment'] = float(legacy_payment.group(1).replace(',', ''))
            data['principal_interest_payment'] = float(legacy_payment.group(2).replace(',', ''))
            data['current_escrow_payment'] = float(legacy_payment.group(3).replace(',', ''))
    if data.get('new_escrow_payment') is None:
        data['new_escrow_payment'] = amount(r'New\s+Escrow\s+Payment')
    if data.get('principal_interest_payment') is None:
        data['principal_interest_payment'] = amount(r'Principal\s*&\s*Interest')
    monthly_line = re.search(r'^\s*Monthly\s+Payment:\s*([^\n]+)', text, re.IGNORECASE | re.MULTILINE)
    monthly_values = [float(item.replace(',', '')) for item in re.findall(r'\$\s*([\d,]+\.\d{2})', monthly_line.group(1))] if monthly_line else []
    if monthly_values:
        data['current_total_payment'] = monthly_values[0]
        data['new_total_payment'] = monthly_values[-1]
    shortage_line = re.search(r'^\s*Shortage:\s*([^\n]+)', text, re.IGNORECASE | re.MULTILINE)
    shortage_values = [float(item.replace(',', '')) for item in re.findall(r'\$\s*([\d,]+\.\d{2})', shortage_line.group(1))] if shortage_line else []
    if shortage_values and shortage_values[-1] > 0:
        data['selected_payment_option'] = 'OPTION_B_WITH_SHORTAGE'

    history_range = re.search(r'Escrow\s+Account\s+Disbursement\s+From\s+([A-Za-z]+\s+\d{4})\s+To\s+([A-Za-z]+\s+\d{4})', text, re.IGNORECASE)
    projection_range = re.search(r'Future\s+Escrow\s+Account\s+Activity\s+For\s+([A-Za-z]+\s+\d{4})\s+To\s+([A-Za-z]+\s+\d{4})', text, re.IGNORECASE)
    if history_range:
        data['history_period_start'] = month_date(history_range.group(1))
        data['history_period_end'] = month_date(history_range.group(2), end=True)
    if projection_range:
        data['projection_period_start'] = month_date(projection_range.group(1))
        data['projection_period_end'] = month_date(projection_range.group(2), end=True)
    if not data.get('history_period_start'):
        explicit_history = re.search(r'Escrow\s+Account\s+History\s+(\d{2}/\d{2}/\d{4})\s+(?:through|to)\s+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if explicit_history:
            data['history_period_start'] = _normalize_date(explicit_history.group(1))
            data['history_period_end'] = _normalize_date(explicit_history.group(2))
    if not data.get('projection_period_start'):
        explicit_projection = re.search(r'Escrow\s+Account\s+Projection\s+(\d{2}/\d{2}/\d{4})\s+(?:through|to)\s+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if explicit_projection:
            data['projection_period_start'] = _normalize_date(explicit_projection.group(1))
            data['projection_period_end'] = _normalize_date(explicit_projection.group(2))
    legacy_activity_range = re.search(r'activity\s+in\s+your\s+Escrow\s+Account\s+from\s+(\d{2}-\d{2}-\d{4})\s+through\s+(\d{2}-\d{2}-\d{4})', text, re.IGNORECASE)
    if legacy_activity_range and not data.get('history_period_start'):
        data['history_period_start'] = _normalize_date(legacy_activity_range.group(1))
        data['history_period_end'] = _normalize_date(legacy_activity_range.group(2))

    history_start = text.lower().find('escrow account activity history for')
    projection_start = text.lower().find('future escrow account activity for')
    history_block = text[history_start:projection_start if projection_start > history_start else len(text)] if history_start >= 0 else ''
    projection_block = text[projection_start:] if projection_start >= 0 else ''
    summary_block = text[text.lower().find('escrow account disbursement from'):history_start] if history_start >= 0 else text
    data['estimated_tax'] = amount(r'Estimated\s+Tax', summary_block)
    data['actual_tax'] = amount(r'Actual\s+Tax', summary_block)
    data['estimated_insurance'] = amount(r'Estimated\s+Insurance', summary_block)
    data['actual_insurance'] = amount(r'Actual\s+Insurance', summary_block)
    data['estimated_total_disbursement'] = amount(r'Estimated\s+Total', summary_block)
    data['actual_total_disbursement'] = amount(r'Actual\s+Total', summary_block)

    projection_summary = text[text.lower().find('escrow account projection'):projection_start] if projection_start >= 0 else text
    data['projected_tax'] = amount(r'(?:COUNTY\s+TAX(?:ES)?|PROPERTY\s+TAX(?:ES)?)', projection_summary)
    data['projected_insurance'] = amount(r'(?:HOMEOWNERS\s+INS(?:URANCE)?|HAZARD\s+INSURANCE)', projection_summary)
    data['projected_total'] = amount(r'(?:Total\s+Annual\s+Taxes\s+And\s+Insurance|Total\s+Escrow)', projection_summary)
    data['projected_monthly_escrow'] = amount(r'New\s+Monthly\s+Escrow\s+Payment', projection_summary)
    if data.get('new_escrow_payment') is None:
        data['new_escrow_payment'] = data.get('projected_monthly_escrow')

    activities = []
    row_pattern = re.compile(r'^\s*(\d{2}/\d{4})\s+(.+?)\s*$', re.MULTILINE)
    for block, phase in ((history_block, 'HISTORICAL'), (projection_block, 'PROJECTED')):
        for match in row_pattern.finditer(block):
            line = match.group(0).strip()
            description_part = match.group(2)
            if description_part.lower().startswith(('totals', 'this amount')):
                continue
            values = [float(value.replace(',', '')) for value in re.findall(r'\$\s*([\d,]+\.\d{2})', line)]
            source_description = re.split(r'\s+\$\s*[\d,]+\.\d{2}', description_part, maxsplit=1)[0].strip()
            item = {
                'activity_date': month_date(match.group(1)),
                'activity_type': category(source_description),
                'source_description': source_description,
                'phase': 'PARTIALLY_PROJECTED' if phase == 'HISTORICAL' and '**' in line else phase,
                'value_status': 'NOT_REMITTED' if '**' in line else ('ACTUAL' if phase == 'HISTORICAL' else 'PROJECTED'),
            }
            if phase == 'HISTORICAL':
                if item['activity_type'] == 'BEGINNING_BALANCE' and len(values) >= 2:
                    item['estimated_balance'], item['actual_balance'] = values[-2:]
                elif len(values) >= 6:
                    item.update(dict(zip(
                        ('estimated_deposit', 'actual_deposit', 'estimated_disbursement', 'actual_disbursement', 'estimated_balance', 'actual_balance'),
                        values[-6:],
                    )))
                    if item['value_status'] == 'NOT_REMITTED':
                        item['actual_deposit'] = item['actual_disbursement'] = item['actual_balance'] = None
            else:
                if item['activity_type'] == 'BEGINNING_BALANCE' and len(values) >= 2:
                    item['estimated_balance'], item['required_balance'] = values[-2:]
                elif len(values) >= 4:
                    item.update(dict(zip(
                        ('estimated_deposit', 'estimated_disbursement', 'estimated_balance', 'required_balance'),
                        values[-4:],
                    )))
            activities.append(item)
    if not activities and legacy_activity_range:
        from datetime import datetime as _dt
        history_end = _dt.strptime(data['history_period_end'], '%Y-%m-%d')
        cursor_year = int(data['history_period_start'][:4])
        previous_month = int(data['history_period_start'][5:7]) - 1
        legacy_rows = []
        for match in re.finditer(r'^\s*(\d{2})\s+([^\n]+)$', text, re.MULTILINE):
            month = int(match.group(1))
            if month < 1 or month > 12:
                continue
            if previous_month and month < previous_month:
                cursor_year += 1
            previous_month = month
            line = match.group(2)
            activity_date = f'{cursor_year:04d}-{month:02d}-01'
            phase = 'HISTORICAL' if (cursor_year, month) <= (history_end.year, history_end.month) else 'PROJECTED'
            description_match = re.search(r'(TRANSFER\s+BAL|COUNTY\s+TAXES|HAZ\s+INSURANC\w*)', line, re.IGNORECASE)
            description = description_match.group(1) if description_match else 'Deposit'
            numbers = [float(value) for value in re.findall(r'(?<!\d)(\d{1,7}(?:\.\d{2}))(?!\d)', line)]
            item = {
                'activity_date': activity_date,
                'activity_type': category(description),
                'source_description': description,
                'phase': phase,
                'value_status': 'ACTUAL' if phase == 'HISTORICAL' else 'PROJECTED',
            }
            if item['activity_type'] in {'PROPERTY_TAX', 'HOMEOWNERS_INSURANCE'} and numbers:
                before_description = line[:description_match.start()] if description_match else line
                disbursement_values = [float(value) for value in re.findall(r'(?<!\d)(\d{1,7}(?:\.\d{2}))(?!\d)', before_description)]
                item['estimated_disbursement'] = disbursement_values[-1] if disbursement_values else numbers[-1]
            elif item['activity_type'] == 'ESCROW_TRANSFER' and numbers:
                item['actual_disbursement'] = numbers[-1]
            elif numbers:
                item['estimated_deposit'] = numbers[0]
                if phase == 'HISTORICAL' and len(numbers) > 1:
                    item['actual_deposit'] = numbers[1]
            legacy_rows.append(item)
        activities = legacy_rows
        projected_dates = [item['activity_date'] for item in activities if item['phase'] == 'PROJECTED']
        if projected_dates:
            data['projection_period_start'] = min(projected_dates)
            last = _dt.strptime(max(projected_dates), '%Y-%m-%d')
            import calendar
            data['projection_period_end'] = last.replace(day=calendar.monthrange(last.year, last.month)[1]).strftime('%Y-%m-%d')
    data['activities'] = activities

    year_source = data.get('statement_date') or data.get('effective_date')
    if year_source:
        data['statement_year'] = int(str(year_source)[:4])
    if data.get('projection_period_start'):
        data['expense_year'] = int(data['projection_period_start'][:4])
    data['period_type'] = 'yearly'
    return {key: value for key, value in data.items() if value is not None}


def detect_period_type(data: Dict[str, Any]) -> str:
    """Detect document period from extracted dates.
    Returns: monthly, quarterly, half_yearly, yearly, other
    """
    start = data.get("period_start") or data.get("statement_date")
    end = data.get("period_end") or data.get("statement_date")
    if not start or not end:
        if data.get("payment_due_date") and data.get("statement_date"):
            return "monthly"
        if data.get("annual_rental_income") is not None or data.get("taxes_paid") is not None:
            return "yearly"
        if data.get("rents_received") is not None:
            return "yearly"
        return "other"

    from datetime import datetime
    for f in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            d1 = datetime.strptime(start, f)
            d2 = datetime.strptime(end, f)
            days = abs((d2 - d1).days)
            if days <= 35:
                return "monthly"
            if days <= 105:
                return "quarterly"
            if days <= 195:
                return "half_yearly"
            return "yearly"
        except (ValueError, TypeError):
            continue
    return "other"


def parse_currency(text: str) -> Optional[float]:
    """Extract first dollar amount from text."""
    match = re.search(r'\$?([\d,]+\.?\d*)', text.replace(',', ''))
    if match:
        try:
            return float(match.group(1).replace(',', ''))
        except ValueError:
            return None
    return None


def parse_percentage(text: str) -> Optional[float]:
    """Extract first percentage from text."""
    match = re.search(r'([\d.]+)\s*%', text)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return None
    return None


# Column-2 labels that bleed into the address line on multi-column statements
_ADDRESS_NOISE = r'\s+(?=(?:Interest|Principal|Escrow|Late|Fees?|Total|Amount|Balance|Payment|Credits?|Activity|Transactions?|Disbursements?|Since|Statement)\b)|\s*\$'

_STREET_SUFFIX = (
    r'(?:St(?:reet)?|Ave(?:nue)?|Dr(?:ive)?|Way|Ct|Court|Ln|Lane|Rd|Road|'
    r'Blvd|Boulevard|Pl(?:ace)?|Cir(?:cle)?|Ter(?:race)?|Pkwy|Parkway|'
    r'Hwy|Highway|Loop|Trail|Trl)'
)


def parse_property_address(text: str) -> Dict[str, Any]:
    """Extract the property address from document text.

    Handles single-line ("Property Address: 456 Oak Ave, Dallas, TX 75201"),
    two-line formats, and multi-column statement layouts where unrelated
    figures share the line ("Property address 6867 Syrah Dr Dublin,CA
    Interest $2,929.16" with the ZIP at the start of the next line).
    """
    # Tolerate the label with or without a space ("Property Address" /
    # "PropertyAddress"). Pick the occurrence whose value starts with a street
    # number — coupon stubs sometimes repeat the label with other text after it.
    data = {}
    line1 = line2 = ''
    address_context = ''
    for mm in re.finditer(r'property\s*address[:\s]*([^\n]*)\n?([^\n]*)?', text, re.IGNORECASE):
        cand1 = (mm.group(1) or '').strip()
        cand2 = (mm.group(2) or '').strip()
        if not cand1:
            cand1, cand2 = cand2, ''
        if re.match(r'^\d', cand1):  # value begins with a house number
            line1, line2 = cand1, cand2
            address_context = text[mm.start():mm.end() + 600]
            break
        if not line1:
            line1, line2 = cand1, cand2
            address_context = text[mm.start():mm.end() + 600]
    if not line1:
        return {}

    # Cut off second-column noise that got merged into the address line
    blob = re.split(_ADDRESS_NOISE, line1)[0].strip(' ,')
    if not blob:
        return {}

    # "City, ST 12345" continuation on the next line — match as a prefix so any
    # trailing table text ("...85258 Paid Last Paid Year to") is ignored.
    csz = re.match(
        r'^([A-Za-z][A-Za-z .\-]*?),?\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b', line2)
    csz_glued = re.match(r'^([A-Za-z]+?)([A-Z]{2})(\d{5})(?:-\d{4})?\b', line2)
    if csz:
        data['property_city'] = csz.group(1).strip()
        data['property_state'] = csz.group(2)
        data['property_zip'] = csz.group(3)
    elif csz_glued:  # run-together "SCOTTSDALEAZ85258"
        data['property_city'] = csz_glued.group(1).strip().title()
        data['property_state'] = csz_glued.group(2)
        data['property_zip'] = csz_glued.group(3)
    # ZIP wrapped to the start of the next line (multi-column layouts)
    elif (zm := re.match(r'^(\d{5})(?:-\d{4})?\b', line2)):
        data['property_zip'] = zm.group(1)

    # Multi-column PDFs can insert payment-table rows between the street and
    # city/state line. Search only the nearby address block so we recover the
    # locality without accidentally using the servicer's mailing address.
    if not all(data.get(key) for key in ('property_city', 'property_state', 'property_zip')):
        nearby_csz = re.search(
            r'^\s*([A-Za-z][A-Za-z .\-]*?),\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b',
            address_context,
            re.MULTILINE,
        )
        if nearby_csz:
            data.setdefault('property_city', nearby_csz.group(1).strip())
            data.setdefault('property_state', nearby_csz.group(2))
            data.setdefault('property_zip', nearby_csz.group(3))

    parts = [p for p in (s.strip() for s in blob.split(',')) if p]
    if not parts:
        return {}
    street = parts[0]

    for tail in parts[1:]:
        sz = re.search(r'\b([A-Z]{2})\b\s*(\d{5})?(?:-\d{4})?', tail)
        if sz and (len(tail) <= 2 or sz.group(2)):
            data['property_state'] = sz.group(1)
            if sz.group(2):
                data['property_zip'] = sz.group(2)
            city = tail[:sz.start()].strip(' ,')
            if city:
                data['property_city'] = city
        else:
            data['property_city'] = tail

    # No comma between street and city: split after the street suffix
    if 'property_city' not in data:
        sm = re.match(r'^(\d+\s+.+?\b' + _STREET_SUFFIX + r'\.?)\s+([A-Za-z][A-Za-z .]*)$',
                      street, re.IGNORECASE)
        if sm:
            street = sm.group(1)
            data['property_city'] = sm.group(2).strip()

    # OCR fix: a scanned statement can render a leading house-number '1' as the
    # letter 'I'/'l' ("I 0308 E San Salvador Dr" → "10308 E San Salvador Dr").
    street = re.sub(r'^[Il]\s?(?=\d)', '1', street)

    data['property_address'] = street
    return data


def parse_borrowers_block(text: str) -> list:
    """Find borrower names from the mailing-address block: one or more
    all-caps name lines directly above 'STREET' / 'CITY ST 12345' lines."""
    lines = [l.strip() for l in text.splitlines()]
    for i, line in enumerate(lines[:-1]):
        if (re.match(r'^\d+\s+[A-Z0-9 .#\-]+$', line)
                and re.match(r'^[A-Z .]+\s[A-Z]{2}\s+\d{5}(-\d{4})?$', lines[i + 1])):
            names = []
            j = i - 1
            while (j >= 0 and len(names) < 4
                   and re.match(r"^[A-Z][A-Z .,'\-]+$", lines[j])
                   and not re.search(r'\d', lines[j])):
                names.insert(0, lines[j])
                j -= 1
            return names
    return []


def _find_amount(text: str, label_patterns: list) -> Optional[float]:
    """Find a dollar amount that follows any of the given label patterns.

    Tolerates 'Label: $1,234.56', 'Label $1,234.56' and table-style layouts
    where the value sits on the same line separated by whitespace or pipes.
    """
    for pat in label_patterns:
        m = re.search(pat + r'[^\d\n%]*?\$?\s*([\d,]+\.\d{1,2}|[\d,]{4,})\b',
                      text, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1).replace(',', ''))
            except ValueError:
                continue
    return None


def _money_values(text: str) -> list:
    values = []
    for raw in re.findall(r'\$?\s*([\d,]+\.\d{2})', text or ''):
        try:
            values.append(float(raw.replace(',', '')))
        except ValueError:
            continue
    return values


_MONTH_NAMES = {m.lower(): i for i, m in enumerate(
    ['January', 'February', 'March', 'April', 'May', 'June', 'July',
     'August', 'September', 'October', 'November', 'December'], 1)}
_MONTH_NAMES.update({m[:3].lower(): i for m, i in list(_MONTH_NAMES.items())})


def _find_date(text: str, label_patterns: list) -> Optional[str]:
    for pat in label_patterns:
        m = re.search(pat + r'[:|\s]*([\d]{1,2}/[\d]{1,4}(?:/[\d]{2,4})?|[\d\-]{6,10})',
                      text, re.IGNORECASE)
        if m:
            return m.group(1)
        # Long-form dates like "May 12, 2026" or "Sept 1 2025" -> MM/DD/YYYY
        m = re.search(
            pat + r'[:|\s]*([A-Za-z]{3,9})\.?\s+(\d{1,2}),?\s+(\d{4})',
            text, re.IGNORECASE)
        if m:
            mon = _MONTH_NAMES.get(m.group(1).lower()[:3])
            if mon:
                return f"{mon:02d}/{int(m.group(2)):02d}/{m.group(3)}"
    return None


def _parse_past_payments_breakdown(text: str) -> Dict[str, Any]:
    """Parse the two-column 'Past Payments Breakdown' table.

    Servicer statements render this as a table with a prior-period column and a
    year-to-date column, e.g.:

        Past Payments Breakdown   As of Last Stmt   Paid Year to Date
        Principal:                $507.12           $3,002.04
        Interest:                 $2,692.72         $16,197.00
        Escrow (Taxes & Ins):     $499.22           $4,464.68
        Total:                    $3,699.06         $23,663.72

    Reading the *first* amount after each label grabs the prior-statement figure,
    not the YTD one — the classic bug. Instead we read the header to learn which
    positional column is YTD, then map each row's amounts to columns by position.
    Handles both "As of Last Stmt / Paid Year to Date" (PennyMac) and
    "Paid Since Last Statement / Paid Year-to-Date" (Chase) orderings, and
    tolerates marketing text interleaved between the rows.
    """
    result: Dict[str, Any] = {}
    m = re.search(r'past\s+payments?\s+breakdown(?P<body>[\s\S]{0,900})', text, re.IGNORECASE)
    if not m:
        return result
    body = m.group('body')
    header = body[:160].lower()

    prior_m = re.search(r'as\s+of\s+last\s+stmt|paid\s+since\s+last\s+statement|since\s+last\s+statement|last\s+statement', header)
    ytd_m = re.search(r'paid\s+year[\s-]*to[\s-]*date|year[\s-]*to[\s-]*date|\bytd\b', header)
    # Default: prior column first, YTD second (the common layout).
    ytd_is_second = True
    if prior_m and ytd_m:
        ytd_is_second = ytd_m.start() > prior_m.start()

    money = r'\$?\s*([\d,]+\.\d{2})'
    rows = (
        ('principal', r'principal'),
        ('interest', r'interest'),
        ('escrow', r'escrow(?:\s*\([^)]*\))?'),
        ('total', r'total'),
    )
    for key, label in rows:
        row = re.search(
            label + r'\s*:?\s*((?:' + money + r'\s*){2,6})',
            body, re.IGNORECASE,
        )
        if not row:
            continue
        amounts = [float(a.replace(',', '')) for a in re.findall(money, row.group(1))]
        if len(amounts) < 2:
            continue
        prior_amt = amounts[0] if ytd_is_second else amounts[1]
        ytd_amt = amounts[1] if ytd_is_second else amounts[0]
        if key == 'principal':
            result['principal_paid_last_statement'] = prior_amt
            result['principal_paid_ytd'] = ytd_amt
        elif key == 'interest':
            result['interest_paid_last_statement'] = prior_amt
            result['interest_paid_ytd'] = ytd_amt
        elif key == 'escrow':
            result['escrow_paid_last_statement'] = prior_amt
            result['escrow_paid_ytd'] = ytd_amt
        elif key == 'total':
            result['total_paid_last_statement'] = prior_amt
            result['total_paid_ytd'] = ytd_amt
    return result


def parse_mortgage_statement(text: str) -> Dict[str, Any]:
    """Extract structured fields from mortgage statement text.

    Targets the schema: property_address, borrowers, account_number,
    original/unpaid principal balance, interest rate, maturity date,
    amount due, principal/interest portions, escrow.
    """
    data = parse_property_address(text)

    borrowers_m = re.search(r'borrowers?[:|]\s*([^\n|]+)', text, re.IGNORECASE)
    if borrowers_m:
        names = [n.strip() for n in re.split(r';|&| and ', borrowers_m.group(1)) if n.strip()]
    else:
        names = parse_borrowers_block(text)
    for i, name in enumerate(names[:4], 1):
        data[f'borrower_{i}'] = name

    # Servicer identity is often presented as a logo, while the legal lender
    # name appears only in the disclosures. Keep both values for matching and
    # display rather than relying on the uploaded filename.
    if re.search(r'JPMorgan\s+Chase\s+Bank\s*,?\s*N\.?A\.?', text, re.IGNORECASE):
        data['lender_name'] = 'JPMorgan Chase Bank, N.A.'
        data['servicer_name'] = 'Chase'
    elif re.search(r'\bCHASE\b', text) and re.search(r'chase\.com/(?:MyMortgage|Statement)', text, re.IGNORECASE):
        data['lender_name'] = 'JPMorgan Chase Bank, N.A.'
        data['servicer_name'] = 'Chase'

    acct_m = re.search(
        r'(?:mortgage\s+)?(?:account|loan)\s+(?:number|no\.?|#)[:|\s]*([\dXx*\-]{4,})',
        text, re.IGNORECASE
    )
    if acct_m:
        data['account_number'] = acct_m.group(1)

    v = _find_amount(text, [
        r'original\s+principal\s+balance',
        r'original\s+loan\s+amount',
    ])
    if v is not None:
        data['original_amount'] = v

    v = _find_amount(text, [
        # "Current principal balance" must win over "Outstanding principal" —
        # servicer statements embed a Form 1098 line ("Box 2: Outstanding
        # mortgage principal") that reports a different (prior year-end) figure.
        r'current\s+principal\s+balance',
        r'unpaid\s+principal\s+balance',
        r'unpaid\s+balance',
        r'outstanding\s+principal(?:\s+balance)?',
        r'outstanding\s+balance',
        r'principal\s+balance',
    ])
    if v is not None:
        data['current_balance'] = v

    # Rate may carry an ARM note, e.g. "Interest rate(Until 10/2032 pmt) 5.12500%"
    rate_m = re.search(r'interest\s+rate\s*(?:\(([^)]*)\))?[:|\s]*([\d.]+)\s*%',
                       text, re.IGNORECASE)
    if rate_m:
        data['interest_rate'] = float(rate_m.group(2))
        if rate_m.group(1):
            data['rate_note'] = rate_m.group(1).strip()
            if 'until' in data['rate_note'].lower():
                data['loan_type'] = 'ARM'

    v = _find_amount(text, [
        r'(?:monthly|regular)\s+payment(?:\s+amount)?',
        r'next\s+payment\s+amount',
        r'amount\s+due',
        r'automatic\s+payment\s+(?:amount|on)',
        r'total\s+payment',
    ])
    if v is None:
        # Statement layouts put the figure on the line below "Amount due"
        m = re.search(r'amount\s+due[^\n]*\n[^\n]*?\$\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
        if m:
            v = float(m.group(1).replace(',', ''))
    if v is not None:
        data['monthly_payment'] = v

    v = _find_amount(text, [
        r'escrow\s*\(taxes\s*and\s*insurance\)',
        r'escrow\s+(?:amount|payment|portion)',
        r'escrow\s*\$',
    ])
    if v is not None:
        data['escrow_amount'] = v
        data['escrow_included'] = True

    # Try to extract property tax separately from escrow breakdown
    v = _find_amount(text, [
        r'property\s+tax(?:es)?\s*(?:paid|amount|escrow)?',
        r'taxes?\s+paid\s+(?:in\s+)?escrow',
        r'taxes?\s+escrow',
        r'taxes:',  # bare "Taxes: $887.47" escrow line on Rocket statements
    ])
    if v is not None:
        data['property_tax_amount'] = v
        data['monthly_property_tax_escrow'] = v

    v = _find_amount(text, [
        r'(?:homeowners?|hazard|property)\s+insurance\s*(?:escrow|premium|amount)?',
        r'insurance\s+escrow',
        r'insurance:',
    ])
    if v is not None:
        data['monthly_insurance_escrow'] = v

    v = _find_amount(text, [
        r'(?:mortgage\s+insurance|pmi)\s*(?:escrow|premium|amount)?',
        r'mortgage\s+insurance:',
        r'pmi:',
    ])
    if v is not None:
        data['monthly_mortgage_insurance'] = v

    component_total = sum(float(data.get(key) or 0) for key in (
        'monthly_property_tax_escrow',
        'monthly_insurance_escrow',
        'monthly_mortgage_insurance',
    ))
    if component_total > 0:
        if not data.get('escrow_amount'):
            data['escrow_amount'] = round(component_total, 2)
        elif round(float(data.get('escrow_amount') or 0), 2) > round(component_total, 2):
            data['monthly_other_escrow'] = round(float(data.get('escrow_amount') or 0) - component_total, 2)
        data['escrow_included'] = True

    # Past Payments Breakdown: read the two-column (prior | year-to-date) table
    # by column position so YTD is never confused with the prior-statement figure.
    breakdown = _parse_past_payments_breakdown(text)
    for field, value in breakdown.items():
        data[field] = value

    # Fallback for single-column "Paid Year to Date" statements (no prior column).
    if 'principal_paid_ytd' not in data:
        ytd_m = re.search(
            r'paid\s+year\s+to\s+date(?P<body>.*?)(?:total\s+paid\s+year\s+to\s+date|escrow\s+disbursements|phone:|page\s+\d)',
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if ytd_m:
            ytd_body = ytd_m.group('body')
            ytd_values = _money_values(ytd_body)
            principal_ytd = _find_amount(ytd_body, [r'principal:'])
            interest_ytd = _find_amount(ytd_body, [r'interest:'])
            escrow_ytd = _find_amount(ytd_body, [r'escrow\s+amount\s*\(taxes\s*&\s*insurance\):', r'escrow:'])
            if principal_ytd is None and len(ytd_values) >= 1:
                principal_ytd = ytd_values[0]
            if interest_ytd is None and len(ytd_values) >= 2:
                interest_ytd = ytd_values[1]
            if escrow_ytd is None and len(ytd_values) >= 3:
                escrow_ytd = ytd_values[2]
            if principal_ytd is not None:
                data['principal_paid_ytd'] = principal_ytd
            if interest_ytd is not None:
                data['interest_paid_ytd'] = interest_ytd
            if escrow_ytd is not None:
                data['escrow_paid_ytd'] = escrow_ytd

    # "principal:" (colon) catches the bare "Principal: $531.46" payment-
    # breakdown label without matching "principal balance:" (no colon there).
    # The current payment's breakdown is listed before any YTD/life-of-loan
    # totals, so the first match is the right one.
    v = _find_amount(text, [
        r'principal\s+portion', r'principal\s+paid', r'principal\s*\$',
        r'principal:',
    ])
    if v is not None:
        data['principal_due'] = v

    v = _find_amount(text, [
        r'interest\s+portion', r'interest\s+paid', r'interest\s*\$',
        r'interest:',
    ])
    if v is not None:
        data['interest_due'] = v

    d = _find_date(text, [r'maturity\s+date'])
    if d:
        data['maturity_date'] = d

    # Loan origination date on servicer "Loan Information" statements.
    d = _find_date(text, [r'origination\s+date'])
    if d:
        data['origination_date'] = d

    # "Original Term 360 months" -> 30 years. Only trust an explicit term line so
    # a stray "N months" (e.g. remaining term) never overrides it.
    term_m = re.search(r'original\s+term\s*[:|\s]*?(\d{2,3})\s*months?', text, re.IGNORECASE)
    if term_m:
        months = int(term_m.group(1))
        data['term_months'] = months
        data['loan_term_years'] = round(months / 12)

    # Servicer identity for statements that only name the lender in fine print.
    if not data.get('lender_name') and re.search(r'pennymac', text, re.IGNORECASE):
        data['lender_name'] = 'PennyMac Loan Services, LLC'
        data['servicer_name'] = 'PennyMac'

    # Online loan-information pages date the snapshot as "Loan Information as of
    # Jun 12, 2026" rather than a labeled "Statement Date". _find_date returns
    # MM/DD/YYYY, keeping the statement_year derivation below correct.
    d = _find_date(text, [r'statement\s+date', r'loan\s+information\s+as\s+of'])
    if d:
        data['statement_date'] = d
        # Extract year from statement date (e.g. "01/15/2024" → 2024, "1/15/24" → 2024)
        parts = re.split(r'[/\-]', d)
        if len(parts) >= 3:
            y = parts[-1]
            if len(y) == 2:
                y = '20' + y
            try:
                data['statement_year'] = int(y)
            except ValueError:
                pass

    d = _find_date(text, [
        r'payment\s+due\s+date', r'next\s+due\s+date', r'due\s+date',
    ])
    if not d:
        m = re.search(r'payment\s+due[^\n]*\n\s*(\d{1,2}/\d{1,2}/\d{2,4})', text, re.IGNORECASE)
        if m:
            d = m.group(1)
    if d:
        data['payment_due_date'] = d

    return data


def parse_tax_return(text: str) -> Dict[str, Any]:
    """Extract rental income data from tax return."""
    data = {}

    # Schedule E rental income
    rental_income_m = re.search(
        r'(?:rents?\s+received|rental\s+income)[:\s]+\$?([\d,]+\.?\d*)',
        text, re.IGNORECASE
    )
    if rental_income_m:
        data['annual_rental_income'] = float(rental_income_m.group(1).replace(',', ''))

    # Depreciation
    depr_m = re.search(
        r'depreciation[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if depr_m:
        data['annual_depreciation'] = float(depr_m.group(1).replace(',', ''))

    # Property taxes (Schedule E line 16)
    taxes_m = re.search(
        r'(?:taxes?|property\s+tax(?:es)?)[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if taxes_m:
        data['taxes_paid'] = float(taxes_m.group(1).replace(',', ''))

    # Total expenses
    expenses_m = re.search(
        r'total\s+expenses[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if expenses_m:
        data['total_expenses'] = float(expenses_m.group(1).replace(',', ''))

    return data


# ── Per-property tax return extraction (Schedule E + Schedule A) ─────────────────
# IRS forms print whole-dollar amounts with a trailing period ("3,585."), which
# lets us tell real values from line numbers/day counts. Schedule E Part I lays
# its three properties out in fixed columns A/B/C; MarkItDown flattens those
# columns, so we read word x-coordinates with pdfplumber and snap each value to
# its column. That recovers the right property even when a row is sparse (e.g.
# only one property has mortgage interest on a given line).

_TR_MONEY = re.compile(r'^\(?-?\$?[\d,]+\.\)?$')


def _tr_value(token: str) -> Optional[float]:
    neg = token.startswith('(') or token.startswith('-')
    t = token.strip('()$').replace(',', '').rstrip('.').lstrip('-')
    if not t:
        return None
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v


def _tr_lines(words, tol: float = 3):
    """Group pdfplumber words into visual lines (same row within `tol` px)."""
    ws = sorted(words, key=lambda w: (w['top'], w['x0']))
    lines = []
    for w in ws:
        for ln in lines:
            if abs(ln['top'] - w['top']) <= tol:
                ln['w'].append(w)
                break
        else:
            lines.append({'top': w['top'], 'w': [w]})
    for ln in lines:
        ln['w'].sort(key=lambda x: x['x0'])
        ln['text'] = ' '.join(x['text'] for x in ln['w'])
    return sorted(lines, key=lambda l: l['top'])


def _tr_money_words(ln):
    out = []
    for w in ln['w']:
        if w['x0'] > 340 and _TR_MONEY.match(w['text']):
            v = _tr_value(w['text'])
            if v is not None:
                out.append((w['x0'], v))
    return out


def _parse_schedule_e_page(page) -> list:
    """Return the rental properties (address + key line items) on one
    Schedule E Part I page, columns recovered from word x-positions."""
    lines = _tr_lines(page.extract_words())

    # Column addresses listed under "1a Physical address of each property"
    addr = {}
    capture = False
    for ln in lines:
        if 'Physical address of each property' in ln['text']:
            capture = True
            continue
        if capture and re.match(r'^1\s*b\b|Type of Property', ln['text']):
            break
        if capture:
            m = re.match(r'^([ABC])\b\s+(.*)$', ln['text'])
            if m:
                addr[m.group(1)] = re.sub(r'^K\s+', '', m.group(2)).strip()
    if not addr:
        return []

    targets = {
        'rents': r'Rents received',
        'advertising': r'\b5\s+Advertising\b',
        'auto_travel': r'\b6\s+Auto\s+and\s+travel\b',
        'cleaning_maintenance': r'\b7\s+Cleaning\s+and\s+maintenance\b',
        'commissions': r'\b8\s+Commissions\b',
        'insurance': r'\b9\s+Insurance\b',
        'legal_professional': r'\b10\s+Legal\s+and\s+other\s+professional\b',
        'management_fees': r'\b11\s+Management\s+fees\b',
        'mortgage_interest': r'Mortgage interest paid to banks',
        'other_interest': r'\b13\s+Other\s+interest\b',
        'repairs': r'\b14\s+Repairs\b',
        'supplies': r'\b15\s+Supplies\b',
        'property_taxes': r'\b16 Taxes\b',
        'utilities': r'\b17\s+Utilities\b',
        'depreciation': r'Depreciation expense',
        'other_expenses': r'\b19\s+Other\b',
        'total_expenses': r'Total expenses',
        # Schedule E lines 2 & 3 — integer day counts, not dollar amounts
        'days_rented':       r'(?:Fair\s+)?[Rr]ental\s+[Dd]ays|(?:2\s+)?(?:Fair\s+)?[Dd]ays\s+rented',
        'personal_use_days': r'[Pp]ersonal\s+[Uu]se\s+[Dd]ays|(?:3\s+)?[Dd]ays\s+(?:of\s+)?personal',
    }
    found, anchor_xs = {}, []
    for ln in lines:
        for key, pat in targets.items():
            if key not in found and re.search(pat, ln['text']):
                mw = _tr_money_words(ln)
                found[key] = mw
                # Fully-populated rows define the column centers
                if key in ('rents', 'total_expenses', 'depreciation'):
                    anchor_xs += [x for x, _ in mw]

    # Cluster x-positions into column centers (left->right = A, B, C)
    anchor_xs.sort()
    clusters = []
    for x in anchor_xs:
        if clusters and x - clusters[-1][-1] <= 25:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    centers = sorted(sum(c) / len(c) for c in clusters)
    cols = ['A', 'B', 'C'][:len(centers)]
    if not centers:
        return []

    def assign(mw):
        out = {}
        for x, v in mw:
            i = min(range(len(centers)), key=lambda i: abs(centers[i] - x))
            out[cols[i]] = v
        return out

    props = []
    for col, address in addr.items():
        vals = {k: assign(found.get(k, [])).get(col, 0.0) or 0.0 for k in targets}
        props.append({
            'address': address,
            'property_kind': 'rental',
            'rents_received': vals['rents'],
            'mortgage_interest': vals['mortgage_interest'],
            'property_taxes': vals['property_taxes'],
            'depreciation': vals['depreciation'],
            'total_expenses': vals['total_expenses'],
            'net_income': round(vals['rents'] - vals['total_expenses'], 2),
            'days_rented':       int(vals['days_rented']),
            'personal_use_days': int(vals['personal_use_days']),
            'expense_breakdown': {
                'advertising': vals['advertising'],
                'auto_travel': vals['auto_travel'],
                'cleaning_maintenance': vals['cleaning_maintenance'],
                'commissions': vals['commissions'],
                'insurance': vals['insurance'],
                'legal_professional': vals['legal_professional'],
                'management_fees': vals['management_fees'],
                'mortgage_interest': vals['mortgage_interest'],
                'other_interest': vals['other_interest'],
                'repairs': vals['repairs'],
                'supplies': vals['supplies'],
                'taxes': vals['property_taxes'],
                'utilities': vals['utilities'],
                'depreciation': vals['depreciation'],
                'other_expenses': vals['other_expenses'],
            },
            'source_refs': {
                'schedule_e': {
                    'page': getattr(page, 'page_number', None),
                    'column': col,
                    'lines': {
                        'address': 'Schedule E Part I line 1a',
                        'days_rented': 'Schedule E Part I line 2',
                        'personal_use_days': 'Schedule E Part I line 3',
                        'rents_received': 'Schedule E Part I line 3',
                        'mortgage_interest': 'Schedule E Part I line 12',
                        'property_taxes': 'Schedule E Part I line 16',
                        'depreciation': 'Schedule E Part I line 18',
                        'total_expenses': 'Schedule E Part I line 20',
                        'net_income': 'Schedule E Part I line 21',
                    },
                },
            },
            'confidence': 0.9 if address and vals['rents'] and vals['total_expenses'] else 0.65,
        })
    return props


def _parse_schedule_a_primary(text: str) -> Optional[dict]:
    """Primary-home mortgage interest (Sch A line 8a) and real estate taxes
    (line 5b). These are the homeowner's own residence, not rentals."""
    # Anchor on the actual Schedule A form header (unique) — not the many
    # incidental "Schedule A" references elsewhere in the return.
    m = re.search(r'SCHEDULE A\s+Itemized Deductions\s+OMB', text, re.IGNORECASE)
    if not m:
        return None
    region = text[m.start():m.start() + 8000]
    interest = re.search(r'\b8a\s+([\d,]+)\.', region)
    taxes = re.search(r'\b5b\s+([\d,]+)\.', region)
    if not interest and not taxes:
        return None
    return {
        'property_kind': 'primary',
        'mortgage_interest': float(interest.group(1).replace(',', '')) if interest else 0.0,
        'property_taxes': float(taxes.group(1).replace(',', '')) if taxes else 0.0,
        'rents_received': 0.0,
        'depreciation': 0.0,
        'total_expenses': 0.0,
        'net_income': 0.0,
        'days_rented': 0,
        'personal_use_days': 0,
    }


def parse_tax_return_properties(filepath: str) -> Dict[str, Any]:
    """Full per-property extraction from a 1040 tax return: every Schedule E
    rental plus the Schedule A primary residence. SSNs and taxpayer names are
    never read into the result — only property figures."""
    import pdfplumber

    full_text = extract_pdf_text(filepath)
    year_m = re.search(r'Schedule\s+E\s*\(Form\s*1040\)\s*(20\d{2})', full_text) \
        or re.search(r'Form\s*1040\s*\((20\d{2})\)', full_text) \
        or re.search(r'\b(20\d{2})\b', full_text)
    tax_year = int(year_m.group(1)) if year_m else None

    schedule1_line5_total = None
    m = re.search(
        r'SCHEDULE\s+1[\s\S]{0,2500}?(?:5\s+)?Rental\s+real\s+estate[\s\S]{0,160}?(\(?-?\$?\s*[\d,]+(?:\.\d{2})?\)?)',
        full_text,
        re.IGNORECASE,
    )
    if m:
        schedule1_line5_total = _tr_value(m.group(1))
    form4562_present = bool(re.search(r'Form\s+4562|Depreciation\s+and\s+Amortization', full_text, re.IGNORECASE))
    depreciation_worksheet_present = bool(re.search(
        r'Depreciation\s+(?:Schedule|Worksheet)|Accumulated\s+Depreciation|Prior\s+Depreciation',
        full_text,
        re.IGNORECASE,
    ))

    properties = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ''
            if 'Physical address of each property' in txt and 'Rents received' in txt:
                properties.extend(_parse_schedule_e_page(page))

    primary = _parse_schedule_a_primary(full_text)
    if primary:
        properties.append(primary)

    rental_total = round(sum(
        p.get('net_income', 0.0)
        for p in properties
        if p.get('property_kind') == 'rental'
    ), 2)
    # The delta is only meaningful as an independent cross-check against a
    # real Schedule 1 figure. When that figure wasn't found, a "delta" against
    # our own Schedule E sum would trivially be ~0 and is misleading — surface
    # a warning instead of a fabricated number.
    validation_delta = None
    schedule1_warning = None
    if schedule1_line5_total is not None:
        validation_delta = round(schedule1_line5_total - rental_total, 2)
    else:
        schedule1_warning = (
            f"Schedule 1 Line 5 not found — using Schedule E sum "
            f"(${rental_total:,.2f}) for cross-check."
        )

    for prop in properties:
        if prop.get('property_kind') != 'rental':
            continue
        prop['schedule1_line5_total'] = schedule1_line5_total
        prop['schedule1_line5_delta'] = validation_delta
        prop['cash_noi'] = round(
            (prop.get('rents_received') or 0)
            - ((prop.get('total_expenses') or 0)
               - (prop.get('mortgage_interest') or 0)
               - (prop.get('depreciation') or 0)),
            2,
        )
        prop['tax_pl'] = prop.get('net_income') or 0.0
        prop['depreciation_detail'] = {
            'form4562_present': form4562_present,
            'depreciation_worksheet_present': depreciation_worksheet_present,
            'current_year_depreciation': prop.get('depreciation') or 0.0,
            'method': None,
            'recovery_period_years': None,
        }
        unresolved = []
        if schedule1_warning:
            unresolved.append(schedule1_warning)
        if prop.get('depreciation'):
            if not form4562_present and not depreciation_worksheet_present:
                unresolved.append(
                    'Depreciation detail missing — upload depreciation schedule for accurate basis tracking.'
                )
            else:
                if not depreciation_worksheet_present:
                    unresolved.append('Depreciation worksheet or asset-level basis not found.')
                if not form4562_present:
                    unresolved.append('Form 4562 not found for depreciation method/recovery-period verification.')
        prop['unresolved_fields'] = unresolved
        prop['source_refs'] = {
            **(prop.get('source_refs') or {}),
            'schedule_1_line_5': {
                'line': 'Schedule 1 line 5',
                'value': schedule1_line5_total,
                'delta_vs_schedule_e_properties': validation_delta,
                'warning': schedule1_warning,
            },
            'form_4562': {
                'present': form4562_present,
                'depreciation_worksheet_present': depreciation_worksheet_present,
            },
        }

    return {
        'tax_year': tax_year,
        'properties': properties,
        'schedule1_line5_total': schedule1_line5_total,
        'schedule1_line5_delta': validation_delta,
        'schedule1_line5_warning': schedule1_warning,
        'form4562_present': form4562_present,
        'depreciation_worksheet_present': depreciation_worksheet_present,
    }


def _money_after(text: str, label: str, window: int = 120,
                 require_dollar: bool = True) -> Optional[float]:
    """Find the first dollar amount that follows a label.

    1098/1099 forms are multi-column; MarkItDown/pdfplumber flatten them so a
    box label and its value land on separate lines, sometimes with form noise
    ("Mortgage Interest Statement") in between. The window spans that gap.

    When *require_dollar* is False the amount may appear without a leading
    '$' sign — some servicer PDFs print Box 2 as a bare number on the next
    line, with the '$' appearing later next to the date in Box 3.
    """
    dollar = r'\$\s*' if require_dollar else r'(?:\$\s*)?'
    m = re.search(
        label + r'[\s\S]{0,%d}?%s([\d,]+\.\d{2})' % (window, dollar),
        text, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1).replace(',', ''))
        except ValueError:
            return None
    return None


def _parse_1098_stacked(text: str) -> Dict[int, Any]:
    """Map a 'Box N:' label list onto the value list that follows it.

    Web year-end statements render every 1098 box label first and then every
    value in the same order (Box 1..Box 10, then $..$..date..$), so a per-label
    window grabs the wrong figure. Pair label and value by position instead.
    A 1099-INT block often follows, so the search is confined to the 1098
    section (between "Form 1098" and the next "Form 1099").
    """
    sec = text
    start = re.search(r'form\s*1098\b', text, re.IGNORECASE)
    if start:
        rest = text[start.end():]
        nxt = re.search(r'form\s*1099\b', rest, re.IGNORECASE)
        sec = rest[:nxt.start()] if nxt else rest
    labels = [int(n) for n in re.findall(r'box\s*(\d+)\s*:', sec, re.IGNORECASE)]
    if not labels:
        return {}
    last = list(re.finditer(r'box\s*\d+\s*:[^\n]*', sec, re.IGNORECASE))[-1].end()
    tokens = re.findall(r'(?:\$\s*)?[\d,]+\.\d{2}|\b\d{2}/\d{2}/\d{4}\b', sec[last:])
    out: Dict[int, Any] = {}
    for box, tok in zip(labels, tokens):
        if '/' in tok:
            out[box] = tok
        else:
            try:
                out[box] = float(tok.replace('$', '').replace(',', '').strip())
            except ValueError:
                pass
    return out


_FORM_1098_MONEY_RE = re.compile(r'\$?\s*(-?[\d,]{1,15}\.\d{2})')
_FORM_1098_DATE_RE = re.compile(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})')
_FORM_1098_INT_RE = re.compile(r'\b(\d{1,3})\b')
_FORM_1098_START = re.compile(r'mortgage\s*interest\s*received\s*from', re.IGNORECASE)
_FORM_1098_TERMINATORS = (
    r'form\s*1098\s*\(keep',
    r'department\s+of\s+the\s+treasury',
    r'www\.irs\.gov/?form\s*1098',
)
_FORM_1098_STREET_SUFFIX = (
    r'WAY|STREET|ST|AVENUE|AVE|DRIVE|DR|LANE|LN|ROAD|RD|COURT|CT|BOULEVARD|BLVD|'
    r'CIRCLE|CIR|PLACE|PL|TERRACE|TER|PARKWAY|PKWY|HIGHWAY|HWY|LOOP|TRAIL|TRL|'
    r'RUN|PASS|POINT|PT|COVE|BEND|SQUARE|SQ|CROSSING|XING|ROW|WALK|PATH'
)


def _isolate_form_1098(text: str) -> str:
    """Return just the Form 1098 region, ignoring coupons / activity pages.

    A servicer year-end packet has several pages; only the IRS 1098 box grid (or
    the "Box N:" snapshot) should be parsed. Each candidate starts at the Box 1
    label and ends at the next form terminator; the candidate that contains the
    most box labels wins.
    """
    starts = [m.start() for m in _FORM_1098_START.finditer(text)]
    if not starts:
        return text
    box_labels = (
        r'outstanding\s*(?:mortgage|principal)', r'mortgage\s*origination\s*date',
        r'refund\s*of\s*overpaid', r'mortgage\s*insurance', r'points\s*paid',
        r'real\s*estate\s*taxes', r'mortgage\s*acquisition\s*date',
        r'address\s*or\s*description\s*of\s*property',
    )
    best, best_score = text, -1
    for start in starts:
        end = len(text)
        for term in _FORM_1098_TERMINATORS:
            tm = re.search(term, text[start:], re.IGNORECASE)
            if tm:
                end = min(end, start + tm.start())
        region = text[start:end]
        score = sum(1 for lbl in box_labels if re.search(lbl, region, re.IGNORECASE))
        if re.search(r'outstanding\s*(?:mortgage|principal)', region, re.IGNORECASE) and _FORM_1098_MONEY_RE.search(region):
            score += 1
        if score > best_score:
            best, best_score = region, score
    return best


def _form_1098_after(region: str, anchor: str, value_re: re.Pattern, window: int = 320):
    """First value matched by *value_re* that appears after *anchor*.

    Value-oriented: we don't stop at the next box label (on grid forms the next
    label sits *between* Box N's label and its value). We simply take the first
    value of the right type after the anchor, within a window — Box N's own value
    is always the first one of its type after its own label.
    """
    m = re.search(anchor, region, re.IGNORECASE)
    if not m:
        return None
    vm = value_re.search(region, m.end(), min(len(region), m.end() + window))
    return vm.group(1) if vm else None


def _form_1098_money(region: str, anchor: str, window: int = 320):
    raw = _form_1098_after(region, anchor, _FORM_1098_MONEY_RE, window)
    if raw is None:
        return None
    try:
        return float(raw.replace(',', ''))
    except ValueError:
        return None


def _form_1098_date(region: str, anchor: str, window: int = 200):
    return _form_1098_after(region, anchor, _FORM_1098_DATE_RE, window)


def _form_1098_integer(region: str, anchor: str, window: int = 60):
    raw = _form_1098_after(region, anchor, _FORM_1098_INT_RE, window)
    try:
        return int(raw) if raw is not None else None
    except ValueError:
        return None


def _form_1098_fragment(region: str, anchor: str, span: int = 180) -> str:
    m = re.search(anchor, region, re.IGNORECASE)
    if not m:
        return region[:span]
    return region[max(0, m.start() - 20):m.end() + span]


def _extract_form_1098_property_address(region: str, full_text: str, data: Dict[str, Any]) -> None:
    """Property address from Box 8 only; fall back to a generic address scan."""
    m = re.search(
        r'address\s*or\s*description\s*of\s*property\s*securing\s*mortgage',
        region, re.IGNORECASE,
    )
    if m:
        after = region[m.end():]
        # Street body is letters/spaces only (no digits or dashes) so the number
        # can't span across neighbouring boxes like "10 Other - Real Estate...".
        street = re.search(
            r'(\d{2,6}\s+[A-Za-z][A-Za-z .]{1,30}?\s+(?:' + _FORM_1098_STREET_SUFFIX + r'))\b',
            after, re.IGNORECASE,
        )
        # The grid interleaves the borrower's city (Box 7) right after the Box 8
        # label, so read the city/state/zip that follows the *property* street,
        # not the first one after the label.
        tail = after[street.end():] if street else after
        csz = re.search(
            r'([A-Za-z][A-Za-z .\-]+?),\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b',
            tail,
        )
        if street:
            data['property_address'] = re.sub(r'\s{2,}', ' ', street.group(1).strip())
        if csz:
            # Drop OCR noise words that grid layouts push in front of the city
            # ("mortgage Paid MANTECA" -> "MANTECA").
            noise = {'paid', 'mortgage', 'other', 'the', 'item', 'real', 'estate',
                     'taxes', 'securing', 'number', 'description', 'address'}
            words = csz.group(1).strip().split()
            while len(words) > 1 and words[0].lower() in noise:
                words.pop(0)
            data['property_city'] = ' '.join(words)
            data['property_state'] = csz.group(2)
            data['property_zip'] = csz.group(3)
    if not data.get('property_address'):
        for key, value in parse_property_address(full_text).items():
            data.setdefault(key, value)


def parse_1098(text: str) -> Dict[str, Any]:
    """Extract data from a Form 1098 Mortgage Interest Statement.

    Maps the IRS boxes onto the app schema:
      Box 1  Mortgage interest received       -> mortgage_interest (annual)
      Box 2  Outstanding mortgage principal   -> current_balance
      Box 3  Mortgage origination date        -> origination_date
      Box 5  Mortgage insurance premiums      -> mortgage_insurance
      Box 6  Points paid                      -> points_paid
      Box 10 Other / Real estate taxes paid   -> property_tax_amount (annual)
    The property address comes from Box 8 (property securing the mortgage),
    not the borrower's mailing address.
    """
    data = {}

    # Calendar / tax year -> drives statement_date and the yearly period.
    # Layouts vary: printed IRS forms say "For calendar year YYYY"; web
    # snapshots say "Tax and Interest Information (YYYY)"; others only show
    # the year in the OMB header box next to "OMB No. 1545-1380".
    # On multi-column statements the year often lands several words after the
    # "For calendar year" label (column noise is interleaved), so allow a gap.
    # "Loan/Escrow Activity YYYY" sections on servicer statements are a
    # reliable fallback. The "(Rev. Month YYYY)" form-revision year always
    # precedes the calendar-year label, so a forward span won't pick it up.
    # Some servicer PDFs have character-level spacing ("c a l e n d a r   y e a r")
    # and the year can land 200+ chars away — widen the window and also try
    # a spacing-tolerant match.
    y = None
    for pat in (
        r'tax\s+and\s+interest\s+information\s*\((\d{4})\)',
        r'calendar\s+year[\s\S]{0,250}?\b(20\d{2})\b',
        r'c\s*\.?\s*a\s*\.?\s*l\s*\.?\s*e\s*\.?\s*n\s*\.?\s*d\s*\.?\s*a\s*\.?\s*r\s+\.?\s*y\s*\.?\s*e\s*\.?\s*a\s*\.?\s*r[\s\S]{0,250}?\b(20\d{2})\b',
        r'(?:loan|escrow)\s+activity\s+(20\d{2})',
        r'1545-1380[\s\S]{0,200}?\b(20\d{2})\b',
    ):
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            y = m.group(1)
            break
    if not y:
        # Last-resort fallback: take the first standalone 20xx year in the
        # first 30 lines of the document (before the instructions section).
        head = '\n'.join(text.splitlines()[:30])
        m = re.search(r'\b(20\d{2})\b', head)
        if m:
            y = m.group(1)
    if y:
        data['tax_year'] = y
        data['statement_year'] = int(y)
        data['statement_date'] = f"12/31/{y}"
        data['period_start'] = f"01/01/{y}"
        data['period_end'] = f"12/31/{y}"

    # ── Structured box extraction ─────────────────────────────────────────────
    # Isolate the Form 1098 region (ignore coupons / activity pages), then read
    # each box by its label anchor + the first value of the expected type that
    # follows it. This is value-oriented, not line-based: it survives OCR that
    # drops the spaces inside a label or wraps the value onto another line, and
    # it handles both the IRS box-grid form and the "Box N:" colon layouts.
    region = _isolate_form_1098(text)

    box1 = _form_1098_money(region, r'mortgage\s*interest\s*received')
    if box1 is not None:
        data['mortgage_interest'] = box1
    box2 = _form_1098_money(region, r'outstanding\s*(?:mortgage|principal)')
    if box2 is not None:
        data['current_balance'] = box2
    box3 = _form_1098_date(region, r'(?:mortgage\s*)?origination\s*date')
    if box3:
        data['origination_date'] = box3
    box4 = _form_1098_money(region, r'refund\s*of\s*overpaid')
    if box4 is not None:
        data['refund_of_overpaid_interest'] = box4
    box5 = _form_1098_money(region, r'mortgage\s*insurance\s*premium')
    if box5:
        data['mortgage_insurance'] = box5
    box6 = _form_1098_money(region, r'points\s*paid')
    if box6:
        data['points_paid'] = box6
    box9 = _form_1098_integer(region, r'number\s*of\s*properties\s*securing')
    if box9 is not None:
        data['property_count'] = box9
    box10 = _form_1098_money(region, r'real\s*estate\s*taxes')
    if box10 is None:
        box10 = _form_1098_money(region, r'(?:^|\n)\s*10\s+other\b')
    if box10 is None:
        box10 = _form_1098_money(region, r'property\s*tax(?:es)?')
    if box10 is not None:
        data['property_tax_amount'] = box10
    box11 = _form_1098_date(region, r'mortgage\s*acquisition\s*date')
    if box11:
        data['mortgage_acquisition_date'] = box11

    # Account number: labelled on its own line, else the first plausible
    # 8-12 digit loan number in the form region.
    m = re.search(
        r'account\s*number(?:\s*\(see\s*instructions\))?[\s:]*\n+\s*([\dXx\*\-]{5,})',
        region, re.IGNORECASE)
    if m:
        data['account_number'] = m.group(1).strip()
    else:
        candidates = [
            c for c in re.findall(r'\b(\d{8,12})\b', region)
            if not c.startswith(('000000', '1545'))
        ]
        if candidates:
            data['account_number'] = candidates[0]

    m = re.search(
        r"recipient'?s/lender'?s\s+name[\s\S]*?(?:telephone no\.?|postal code[^\n]*)\s*\n\s*([^\n]+)",
        text, re.IGNORECASE)
    if m:
        name = m.group(1).strip()
        # Reject lines that are really wrapped form boilerplate, not a name
        boilerplate = ('telephone', 'limits based', 'deductible', 'postal',
                       'caution', 'secured property', '1098')
        if name and not any(w in name.lower() for w in boilerplate):
            data['lender_name'] = name

    # Box 8 - property securing the mortgage. Read ONLY the region that follows
    # the Box 8 label so the borrower's mailing address (Box 7) and statement
    # noise ("Credits since last statement") can never leak in.
    _extract_form_1098_property_address(region, text, data)

    # ── Validation: never silently drop Box 2 ────────────────────────────────
    # A real 1098 always reports outstanding principal. If Box 1 parsed but
    # Box 2 did not, the extraction is wrong (not the document) — log the OCR
    # fragment and fail loudly rather than returning a null the UI would ask the
    # user to fill in by hand.
    if data.get('mortgage_interest') is not None and data.get('current_balance') is None:
        fragment = _form_1098_fragment(region, r'outstanding\s*(?:mortgage|principal)')
        logger.error(
            "Form 1098 Box 2 (outstanding mortgage principal) missing though Box 1 "
            "was extracted. OCR fragment near Box 2: %r", fragment,
        )
        raise ValueError(
            "Form 1098 parse failed: Box 2 (outstanding mortgage principal) could "
            f"not be read. OCR fragment near Box 2: {fragment!r}"
        )

    # Borrower name(s) - capture lines after the wrapped label, before the street
    m = re.search(
        r"payer'?s/borrower'?s\s+name[\s\S]*?postal code\s*\n([\s\S]*?)\n\s*\d",
        text, re.IGNORECASE)
    if m:
        names = [l.strip() for l in m.group(1).splitlines()
                 if l.strip() and not re.search(r'\d', l)]
        for i, nm in enumerate(names[:4], 1):
            data[f'borrower_{i}'] = nm

    return data


def parse_1099(text: str) -> Dict[str, Any]:
    """Extract data from 1099 form."""
    data = {}
    # Box 1 - rents
    box1_m = re.search(r'(?:box\s*1|rents)[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE)
    if box1_m:
        data['rents_received'] = float(box1_m.group(1).replace(',', ''))
    return data


def parse_bank_statement(text: str) -> Dict[str, Any]:
    """Extract data from a bank statement."""
    data = parse_property_address(text)

    # Bank name
    bank_m = re.search(r'([A-Za-z\s]+(?:bank|credit union|financial|savings))',
                       text, re.IGNORECASE)
    if bank_m:
        data['bank_name'] = bank_m.group(1).strip()

    # Account number
    acct_m = re.search(
        r'(?:account|acct)\s+(?:number|no\.?|#)[:|\s]*([\dXx*\-]{4,})',
        text, re.IGNORECASE
    )
    if acct_m:
        data['account_number'] = acct_m.group(1)

    # Statement period
    period_m = re.search(
        r'(?:statement\s+)?(?:period|for\s+the\s+month|for\s+period)[\s:]+\
([A-Za-z]+\s+\d{1,2},?\s*\d{4})\s*(?:through|to|-)\s*([A-Za-z]+\s+\d{1,2},?\s*\d{4})',
        text, re.IGNORECASE
    )
    if not period_m:
        period_m = re.search(
            r'(?:statement\s+)?(?:period|for\s+the\s+month|for\s+period)[\s:]+\
(\d{1,2}/\d{1,2}/\d{2,4})\s*(?:through|to|-)\s*(\d{1,2}/\d{1,2}/\d{2,4})',
            text, re.IGNORECASE
        )
    if period_m:
        data['period_start'] = period_m.group(1)
        data['period_end'] = period_m.group(2)

    # Beginning balance
    v = _find_amount(text, [
        r'beginning\s+balance', r'starting\s+balance', r'opening\s+balance',
    ])
    if v is not None:
        data['beginning_balance'] = v

    # Ending balance
    v = _find_amount(text, [
        r'ending\s+balance', r'closing\s+balance',
    ])
    if v is not None:
        data['ending_balance'] = v

    # Total deposits / credits
    v = _find_amount(text, [
        r'total\s+deposits?', r'total\s+credits?', r'deposits?\s+and\s+credits?',
    ])
    if v is not None:
        data['total_deposits'] = v

    # Total withdrawals / debits
    v = _find_amount(text, [
        r'total\s+withdrawals?', r'total\s+debits?', r'withdrawals?\s+and\s+debits?',
    ])
    if v is not None:
        data['total_withdrawals'] = v

    # No statement_date — use period_end
    if not data.get('statement_date') and data.get('period_end'):
        data['statement_date'] = data['period_end']

    return data


def parse_property_tax(text: str) -> Dict[str, Any]:
    """Extract property tax details from annual/half-yearly tax statements."""
    data = parse_property_address(text)

    # Tax year. The county's own "TAX YEAR: YYYY" stub is authoritative; a
    # fiscal-range title like "2023 - 2024 PROPERTY TAX BILL" otherwise gets
    # mis-read as the later year, colliding bills from adjacent years.
    year_m = re.search(r'tax\s+year[:\s]+(\d{4})', text, re.IGNORECASE)
    if not year_m:
        # Fiscal-range title: use the starting year ("2023 - 2024" -> 2023).
        year_m = re.search(
            r'(\d{4})\s*[-–]\s*\d{4}\s+property\s+tax', text, re.IGNORECASE)
    if not year_m:
        year_m = re.search(r'(\d{4})\s+property\s+tax', text, re.IGNORECASE)
    if not year_m:
        year_m = re.search(r'property\s+tax\s+\w+\s+(\d{4})', text, re.IGNORECASE)
    if not year_m:
        year_m = re.search(r'(\d{4})\s+assessed\s+taxes', text, re.IGNORECASE)
    if year_m:
        data['tax_year'] = year_m.group(1)
        data['statement_date'] = f"01/01/{year_m.group(1)}"
        data['statement_year'] = int(year_m.group(1))
        data['period_start'] = f"01/01/{year_m.group(1)}"
        data['period_end'] = f"12/31/{year_m.group(1)}"

    # Total taxes
    v = _find_amount(text, [
        r'total\s+\d{4}\s+property\s+taxes?',
        r'total\s+assessed\s+taxes?',
        r'your\s+total\s+\d{4}\s+property\s+taxes?',
        r'your\s+total\s+property\s+taxes?',
        # Alameda County: "Ad Valorem Tax plus Special Assessments ... $X"
        r'ad\s+valorem\s+tax\s+plus\s+special\s+assessments',
    ])
    # "Total Amount Billed" header with amount on next line (Alameda County portal)
    if v is None:
        m = re.search(
            r'total\s+amount\s+billed[^\n]*\n[^\n]*?\$?\s*([\d,]+\.\d{2})',
            text, re.IGNORECASE)
        if m:
            try:
                v = float(m.group(1).replace(',', ''))
            except ValueError:
                pass
    # County installment bills (e.g. San Joaquin CA) show no single "total"
    # line — the annual tax is the sum of the 1st and 2nd installments.
    if v is None:
        installments = re.findall(
            r'(?:1st|2nd|first|second)\s+installment\s*\$?\s*([\d,]+\.\d{2})',
            text, re.IGNORECASE)
        if installments:
            v = round(sum(float(x.replace(',', '')) for x in installments), 2)
    if v is not None:
        data['taxes_paid'] = v
        data['property_tax_amount'] = v

    # Parcel / APN
    parcel_m = re.search(r'parcel[#\s]+([\d\-]+)', text, re.IGNORECASE)
    if parcel_m:
        data['parcel_number'] = parcel_m.group(1)

    # Assessed values
    v = _find_amount(text, [r'limited\s+value'])
    if v is not None:
        data['limited_value'] = v
    v = _find_amount(text, [r'full\s+cash\s+value', r'secondary\s+full\s+cash'])
    if v is not None:
        data['full_cash_value'] = v

    # Half-year amounts from payment stubs
    halves = re.findall(r'(?:first|second)\s+half[^$]*\$?\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
    if halves:
        data['first_half'] = float(halves[0].replace(',', ''))
        if len(halves) > 1:
            data['second_half'] = float(halves[1].replace(',', ''))

    # Previous year total for comparison
    v = _find_amount(text, [r'previous\s+year\s+total'])
    if v is not None:
        data['previous_year_taxes'] = v

    return data


def parse_insurance_declaration(text: str) -> Dict[str, Any]:
    """Extract the annual premium and coverage period from an insurance declaration."""
    data: Dict[str, Any] = parse_property_address(text)

    effective = re.search(
        r'(?:policy\s+)?effective\s+date\s*:?[\s$]*([A-Za-z]+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})',
        text,
        re.IGNORECASE,
    )
    expiration = re.search(
        r'(?:policy\s+)?expiration\s+date\s*:?[\s$]*([A-Za-z]+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})',
        text,
        re.IGNORECASE,
    )
    period = re.search(
        r'policy\s+period\s*:?[\s$]*([A-Za-z]+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})\s+(?:to|through|-)\s+([A-Za-z]+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})',
        text,
        re.IGNORECASE,
    )
    if period:
        data['period_start'] = _normalize_date(period.group(1))
        data['period_end'] = _normalize_date(period.group(2))
    else:
        if effective:
            data['period_start'] = _normalize_date(effective.group(1))
        if expiration:
            data['period_end'] = _normalize_date(expiration.group(1))

    premium = _find_amount(text, [
        r'total\s+(?:annual\s+)?policy\s+premium',
        r'annual\s+(?:homeowners?\s+|hazard\s+)?insurance\s+premium',
        r'total\s+premium',
        r'policy\s+premium',
    ])
    if premium is None:
        monthly = _find_amount(text, [r'monthly\s+(?:insurance\s+)?premium'])
        if monthly is not None:
            premium = round(monthly * 12, 2)
    if premium is not None:
        data['annual_insurance'] = round(premium, 2)

    year_source = data.get('period_start') or data.get('period_end')
    if not year_source:
        statement = re.search(
            r'(?:statement|issue)\s+date\s*:?[\s$]*([A-Za-z]+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})',
            text,
            re.IGNORECASE,
        )
        if statement:
            data['statement_date'] = _normalize_date(statement.group(1))
            year_source = data['statement_date']
    if year_source:
        year_match = re.search(r'(?:19|20)\d{2}', str(year_source))
        if year_match:
            data['statement_year'] = int(year_match.group(0))
    data['period_type'] = 'yearly'
    return data


def parse_loan_disclosure(text: str) -> Dict[str, Any]:
    """Extract opening loan terms from a Loan Estimate or refinance disclosure.

    A disclosure establishes debt terms. It does not report YTD activity or a
    current verified statement balance, so those fields remain the domain of
    1098 forms and monthly mortgage statements.
    """
    is_closing_disclosure = bool(re.search(
        r'^\s*closing\s+disclosure\s*$',
        text,
        re.IGNORECASE | re.MULTILINE,
    ))
    data = parse_closing_statement(text) if is_closing_disclosure else parse_property_address(text)
    if re.search(r'^\s*loan\s+estimate\s*$', text, re.IGNORECASE | re.MULTILINE):
        data['document_type'] = 'LOAN_ESTIMATE'
        data['classification_confidence'] = 0.99

    # CFPB transaction tables often use "Property" without a colon and may
    # include checkbox artifacts after the street on the same line.
    property_m = re.search(
        r'^\s*Property\s*:?[ \t]*'
        r'(\d+[^\n]*?\b' + _STREET_SUFFIX + r'\.?)[^\n]*\n\s*'
        r'([^,\n]+),\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b',
        text,
        re.IGNORECASE | re.MULTILINE,
    )
    if property_m:
        data['property_address'] = property_m.group(1).strip()
        data['property_city'] = property_m.group(2).strip()
        data['property_state'] = property_m.group(3).upper()
        data['property_zip'] = property_m.group(4)

    # Full identifier, including hyphenated suffixes. The older closing parser
    # intentionally accepted numeric IDs only and truncated this value.
    account_candidates = []
    for account_line in re.findall(r'^.*loan\s*id\s*#?.*$', text, re.IGNORECASE | re.MULTILINE):
        tail = re.split(r'loan\s*id\s*#?', account_line, maxsplit=1, flags=re.IGNORECASE)[-1]
        account_m = re.match(r'\s*([A-Z0-9-]+(?:\s+(?=\d)\d+)*)', tail, re.IGNORECASE)
        if account_m:
            account_candidates.append(re.sub(r'\s+', '', account_m.group(1)))
    if account_candidates:
        data['account_number'] = max(account_candidates, key=len)
        data['loan_id'] = data['account_number']

    purpose_m = re.search(r'\bPurpose\s+(Purchase|Refinance|Construction|Home\s+Equity)\b', text, re.IGNORECASE)
    if purpose_m:
        data['loan_purpose'] = purpose_m.group(1).strip().title()
        data['transaction_purpose'] = data['loan_purpose'].upper()

    product_m = re.search(r'\bLoan\s+Type\b[^\n]*?\b(Conventional|FHA|VA|USDA)\b', text, re.IGNORECASE)
    if product_m:
        data['loan_product'] = product_m.group(1).upper() if product_m.group(1).upper() in {'FHA', 'VA', 'USDA'} else product_m.group(1).title()
    elif re.search(r'C\s*o\s*n\s*v\s*e\s*n\s*t\s*i\s*o\s*n\s*a\s*l', text, re.IGNORECASE):
        data['loan_product'] = 'Conventional'

    # Prefer the transaction-information lender. Some PDF extractors place the
    # lender label on the following line after both settlement-agent and lender
    # names have been flattened together.
    transaction_lender = re.search(
        r'\bLender[ \t]+(.+?)[ \t]+MIC\s*#',
        text,
        re.IGNORECASE,
    )
    lender_m = re.search(r'^\s*Lender[ \t]*:?[ \t]+(?!Credits\b)([^\n]+)', text, re.IGNORECASE | re.MULTILINE)
    lender = transaction_lender.group(1).strip() if transaction_lender else (lender_m.group(1).strip() if lender_m else '')
    if not lender:
        stacked_lender = re.search(r'^([^\n]+)\n\s*Lender\s*$', text, re.IGNORECASE | re.MULTILINE)
        previous_line = stacked_lender.group(1).strip() if stacked_lender else ''
        split_agent = re.search(
            r'(?:Inc\.?|LLC|Ltd\.?|Company|Corporation)\s+(.+)$',
            previous_line,
            re.IGNORECASE,
        )
        if split_agent:
            lender = split_agent.group(1).strip()
    if lender and not lender.lower().startswith('credits'):
        data['lender_name'] = lender.rstrip('.')

    # Loan amount
    loan_m = re.search(
        r'loan\s+amount[:\s]+\$?([\d,]+\.?\d*)', text, re.IGNORECASE
    )
    if not loan_m:
        loan_m = re.search(
            r'\$?([\d,]+\.?\d*)\s*\n?\s*Loan\s+Amount\b',
            text,
            re.IGNORECASE,
        )
    if loan_m:
        original_amount = float(loan_m.group(1).replace(',', ''))
        if original_amount >= 10000:
            data['original_amount'] = original_amount
            data['current_balance'] = original_amount
            data['current_balance_source'] = 'loan_disclosure_initial_balance'
            data['current_balance_verified'] = False

    # Rate
    rate_m = re.search(r'interest\s+rate[:\s]+([\d.]+)\s*%', text, re.IGNORECASE)
    if not rate_m:
        # Some CFPB PDFs flatten the value before its label:
        # "4.125% NO" followed by "Interest Rate".
        rate_m = re.search(
            r'([\d.]+)\s*%\s*(?:NO|YES)?\s*\n?\s*Interest\s+Rate\b',
            text,
            re.IGNORECASE,
        )
    if rate_m:
        data['interest_rate'] = float(rate_m.group(1))

    # Loan term
    term_m = re.search(r'loan\s+term[:\s]+(\d+)\s*years?', text, re.IGNORECASE)
    if not term_m:
        term_m = re.search(r'(\d+)\s*years?\s*\n\s*loan\s+term\b', text, re.IGNORECASE)
    if term_m:
        data['loan_term_years'] = int(term_m.group(1))

    # Loan type
    if re.search(r'\bARM\b|\badjustable\b', text, re.IGNORECASE):
        data['loan_type'] = 'ARM'
    elif re.search(r'\bfixed\b', text, re.IGNORECASE):
        data['loan_type'] = 'FIXED'

    # Monthly payment
    payment_m = re.search(
        r'(?:principal\s*&\s*interest|monthly\s+payment)[:\s]+\$?([\d,]+\.?\d*)',
        text, re.IGNORECASE
    )
    if payment_m:
        data['monthly_payment'] = float(payment_m.group(1).replace(',', ''))

    if data.get('loan_type') == 'ARM':
        if re.search(r'beginning\s+of\s+121st\s+month', text, re.IGNORECASE):
            data['arm_first_change_month'] = 121
            data['arm_product'] = '10/6 ARM'
        frequency_m = re.search(r'every\s+(\d+)\s+months?\s+after\s+first\s+change', text, re.IGNORECASE)
        if frequency_m:
            data['arm_adjustment_frequency_months'] = int(frequency_m.group(1))
        margin_m = re.search(r'index\s*\+\s*margin\s+.+?\+\s*([\d.]+)\s*%', text, re.IGNORECASE)
        if margin_m:
            data['arm_margin'] = float(margin_m.group(1))
        limits_m = re.search(r'minimum/maximum\s+interest\s+rate\s+([\d.]+)%\s*/\s*([\d.]+)%', text, re.IGNORECASE)
        if limits_m:
            data['minimum_interest_rate'] = float(limits_m.group(1))
            data['maximum_interest_rate'] = float(limits_m.group(2))

    issued_m = re.search(r'date\s+issued\s+(\d{1,2}/\d{1,2}/\d{2,4})', text, re.IGNORECASE)
    if issued_m:
        data['issued_date'] = _normalize_date(issued_m.group(1))

    sale_price = _find_amount(text, [r'sale\s+price'])
    if sale_price is not None:
        data['purchase_price'] = sale_price
        data['sale_price'] = sale_price
    estimated_closing = _find_amount(text, [r'estimated\s+closing\s+costs'])
    if estimated_closing is None:
        estimated_closing_m = re.search(
            r'\$\s*([\d,]+(?:\.\d{1,2})?)\s+[^\n]*(?:\n\s*)?estimated\s+closing\s+costs',
            text,
            re.IGNORECASE,
        )
        if estimated_closing_m:
            estimated_closing = float(estimated_closing_m.group(1).replace(',', ''))
    if estimated_closing is not None:
        data['closing_costs'] = estimated_closing
    estimated_cash = _find_amount(text, [r'estimated\s+cash\s+to\s+close'])
    if estimated_cash is None:
        estimated_cash_m = re.search(
            r'\$\s*([\d,]+(?:\.\d{1,2})?)\s+[^\n]*(?:\n\s*)?estimated\s+cash\s+to\s+close',
            text,
            re.IGNORECASE,
        )
        if estimated_cash_m:
            estimated_cash = float(estimated_cash_m.group(1).replace(',', ''))
    if estimated_cash is not None:
        data['cash_to_close'] = estimated_cash

    # Derive the down payment only when financing and cash-to-close math
    # independently reconcile to the same value.
    if not data.get('down_payment') and data.get('purchase_price') and data.get('original_amount'):
        financing_down = round(float(data['purchase_price']) - float(data['original_amount']), 2)
        cash_down = None
        if data.get('cash_to_close') is not None and data.get('closing_costs') is not None:
            cash_down = round(float(data['cash_to_close']) - float(data['closing_costs']), 2)
        if financing_down >= 0 and cash_down is not None and abs(financing_down - cash_down) <= 1:
            data['down_payment'] = financing_down
            data['down_payment_source'] = 'purchase_price_minus_loan_amount_reconciled_to_cash_to_close'

    property_costs = _find_amount(text, [r'estimated\s+taxes\s*,?\s*insurance\s*&\s*assessments'])
    if property_costs is None:
        property_costs_m = re.search(
            r'estimated\s+taxes\s*,?\s*insurance(?:\s*&\s*assessments)?\s*\$\s*([\d,]+(?:\.\d{1,2})?)',
            text,
            re.IGNORECASE,
        )
        if property_costs_m:
            property_costs = float(property_costs_m.group(1).replace(',', ''))
    if property_costs is not None:
        data['estimated_non_escrow_property_costs_monthly'] = property_costs
    if re.search(r'estimated\s+escrow\s+\+?\s*0\b', text, re.IGNORECASE):
        data['escrow_included'] = False

    closing_m = re.search(r'closing\s+date\s*:?\s*(\d{1,2}/\d{1,2}/\d{2,4})', text, re.IGNORECASE)
    if closing_m:
        data['origination_date'] = _normalize_date(closing_m.group(1))

    if data.get('monthly_payment') and data.get('estimated_total_monthly_payment') is None:
        data['estimated_total_monthly_payment'] = data['monthly_payment']

    return data


def _normalize_date(s: str) -> Optional[str]:
    """Convert human-readable date strings to YYYY-MM-DD; return as-is if not parseable."""
    from datetime import datetime
    for fmt in ("%B %d, %Y", "%B %d %Y", "%b %d, %Y", "%b %d %Y",
                "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s.strip()


def _parse_amount(s: str) -> Optional[float]:
    """Parse a dollar amount string with or without cents."""
    try:
        return float(re.sub(r'[,$\s]', '', s))
    except (ValueError, TypeError):
        return None


def _line_after_label_amount(
    text: str,
    label_pattern: str,
    *,
    min_value: Optional[float] = None,
    prefer_last_on_line: bool = False,
) -> Optional[float]:
    m = re.search(label_pattern, text, re.IGNORECASE | re.MULTILINE)
    if not m:
        return None
    line_start = text.rfind('\n', 0, m.start()) + 1
    line_end = text.find('\n', m.end())
    if line_end == -1:
        line_end = len(text)
    line_rest = text[m.end():line_end]
    rest = line_rest if line_rest.strip() else text[m.end(): m.end() + 160]

    # Prefer comma-formatted money first. This avoids reading dates or rates
    # such as "4/12" or "4.12%" as a mortgage principal.
    matches = re.findall(r'\$?\s*(\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?)', rest)
    if not matches and min_value is None:
        matches = re.findall(r'\$?\s*(\d+(?:\.\d{1,2})?)', rest)
    amounts = [_parse_amount(match) for match in matches]
    amounts = [amount for amount in amounts if amount is not None and (min_value is None or amount >= min_value)]
    if not amounts:
        return None
    return amounts[-1] if prefer_last_on_line else amounts[0]


def _extract_alta_settlement_calculations(text: str) -> Dict[str, Any]:
    """Extract ALTA settlement totals and purchase-price candidates.

    MarkItDown preserves these forms well enough to capture the final
    balancing totals, while visual debit/credit columns can flatten. Keep
    settlement math backend-owned and expose explicit setup-review candidates
    instead of making the React form infer financial meaning.
    """
    data: Dict[str, Any] = {}
    compact = re.sub(r'\s+', ' ', text)
    lines = [line.strip() for line in text.splitlines()]

    payoff_match = re.search(
        r'\bPayoff\(s\)\s+(?:to\s+)?(.+?)\s+\$\s*([\d,]+(?:\.\d{1,2})?)\b',
        compact,
        re.IGNORECASE,
    )
    if payoff_match:
        payoff_lender = re.sub(r'\s+', ' ', payoff_match.group(1)).strip(' :-')
        payoff_amount = _parse_amount(payoff_match.group(2))
        if payoff_lender:
            data["prior_loan_payoff_lender"] = payoff_lender
        if payoff_amount is not None:
            data["prior_loan_payoff_amount"] = payoff_amount

    def numeric_line_values(start_index: int, stop_patterns: tuple[str, ...] = ()) -> list[float]:
        values = []
        for line in lines[start_index:]:
            if any(re.search(pattern, line, re.IGNORECASE) for pattern in stop_patterns):
                break
            if re.fullmatch(r'\$?\s*[\d,]+(?:\.\d{1,2})?', line):
                amount = _parse_amount(line)
                if amount is not None:
                    values.append(amount)
        return values

    if "Final Settlement Statement" in text and "Sale Price" in text and "Loan Amount" in text:
        debit_index = next((index for index, line in enumerate(lines) if re.fullmatch(r'Debit', line, re.IGNORECASE)), -1)
        if debit_index >= 0:
            leading_amounts = numeric_line_values(debit_index + 1, (r'^Page\s+1\b', r'^Description$'))
            if len(leading_amounts) >= 3:
                data.setdefault("sale_price", leading_amounts[0])
                data.setdefault("deposit", leading_amounts[1])
                data.setdefault("loan_amount", leading_amounts[2])

    due_match = re.search(
        r'Subtotals\s+'
        r'([\d,]+(?:\.\d{1,2})?)\s+'
        r'([\d,]+(?:\.\d{1,2})?)\s+'
        r'Due\s+To\s+Buyer\s+'
        r'([\d,]+(?:\.\d{1,2})?)\s+'
        r'([\d,]+(?:\.\d{1,2})?)\s+'
        r'Totals\s+([\d,]+(?:\.\d{1,2})?)\s+'
        r'([\d,]+(?:\.\d{1,2})?)\s+'
        r'([\d,]+(?:\.\d{1,2})?)',
        compact,
        re.IGNORECASE,
    )
    if due_match:
        debit_subtotal = _parse_amount(due_match.group(1))
        credit_subtotal = _parse_amount(due_match.group(2))
        due_to_buyer = _parse_amount(due_match.group(3))
        debit_total = _parse_amount(due_match.group(6))
        credit_total = _parse_amount(due_match.group(7))
        data.update({
            "settlement_debit_subtotal": debit_subtotal,
            "settlement_credit_subtotal": credit_subtotal,
            "settlement_due_to_buyer": due_to_buyer,
            "settlement_debit_total": debit_total,
            "settlement_credit_total": credit_total,
            "settlement_total_amount": credit_total or debit_total or credit_subtotal,
            "settlement_total_source": "document_totals_last_line",
        })
    elif "Subtotals" in text and "Due To Buyer" in text and "Totals" in text:
        subtotal_index = next((index for index, line in enumerate(lines) if re.fullmatch(r'Subtotals', line, re.IGNORECASE)), -1)
        if subtotal_index >= 0:
            totals = numeric_line_values(subtotal_index + 1, (r'^Copyright', r'^Page\s+\d+\b'))
            if len(totals) >= 3:
                subtotal = totals[-5] if len(totals) >= 5 else totals[0]
                due_to_buyer = totals[-4] if len(totals) >= 5 else totals[1]
                settlement_total = totals[-1]
                data.update({
                    "settlement_debit_subtotal": subtotal,
                    "settlement_credit_subtotal": subtotal,
                    "settlement_due_to_buyer": due_to_buyer,
                    "settlement_debit_total": settlement_total,
                    "settlement_credit_total": settlement_total,
                    "settlement_total_amount": settlement_total,
                    "settlement_total_source": "document_totals_last_line",
                })

    # Settlement forms can flatten the final balancing row to a single
    # ``TOTALS`` line without the preceding labels. On settlement/ALTA/HUD
    # documents, the final amount in the last TOTALS row is the authoritative
    # final settlement total.
    if not data.get("settlement_total_amount") and re.search(
        r'final\s+settlement\s+statement|alta\s+settlement\s+statement|hud-?1\s+settlement\s+statement',
        text,
        re.IGNORECASE,
    ):
        totals_indexes = [
            index for index, line in enumerate(lines)
            if re.match(r'^totals\b', line, re.IGNORECASE)
        ]
        if totals_indexes:
            totals_index = totals_indexes[-1]
            totals_line = lines[totals_index]
            total_values = [
                amount for amount in (
                    _parse_amount(value)
                    for value in re.findall(r'\$?\s*(\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?)', totals_line)
                )
                if amount is not None
            ]
            if not total_values:
                total_values = numeric_line_values(totals_index + 1, (r'^Copyright', r'^Page\s+\d+\b'))[:4]
            if total_values:
                settlement_total = total_values[-1]
                data["settlement_total_amount"] = settlement_total
                data["settlement_total_source"] = "document_totals_last_line"
                if len(total_values) >= 2:
                    data.setdefault("settlement_debit_total", total_values[-2])
                    data.setdefault("settlement_credit_total", settlement_total)

    # Layout-preserving PDF extraction commonly keeps the final balancing row
    # on one line. Capture each accounting column explicitly. These amounts
    # are retained for audit and are not treated as purchase price or costs.
    totals_lines = [line for line in lines if re.match(r'^TOTALS\b', line, re.IGNORECASE)]
    if totals_lines:
        total_values = [
            amount for amount in (
                _parse_amount(value)
                for value in re.findall(r'([\d,]+\.\d{2})', totals_lines[-1])
            )
            if amount is not None
        ]
        if len(total_values) >= 2:
            data['settlement_debit_total'] = total_values[-2]
            data['settlement_credit_total'] = total_values[-1]
            data['settlement_total_amount'] = total_values[-1]
            data['settlement_total_source'] = 'settlement_credit_total'

    balance_due_match = re.search(
        r'(?:Balance\s+)?Due\s+FROM\s+Buyer[^\n\d]*([\d,]+\.\d{2})',
        text,
        re.IGNORECASE,
    )
    if not balance_due_match:
        balance_due_match = re.search(
            r'Due\s+From\s+Borrower[^\n\d]*([\d,]+\.\d{2})',
            text,
            re.IGNORECASE,
        )
    if balance_due_match:
        balance_due = _parse_amount(balance_due_match.group(1))
        data['cash_to_close'] = balance_due
        data['balance_due_from_buyer'] = balance_due
        data['total_due_from_borrower_cash'] = balance_due

    important_fields = [
        ("sale_price", "Sale Price", r'\bSale\s+Price\b'),
        ("deposit", "Deposit", r'\bDeposit\s*:'),
        ("loan_amount", "Loan Amount", r'\bLoan\s+Amount\b'),
        ("initial_deposit_retained_by_seller", "Initial Deposit retained by Seller", r'\bInitial\s+Deposit\s+retained\s+by\s+Seller\b'),
        ("preferred_incentive", "Preferred Incentive", r'\bPreferred\s+Incentive\b'),
        ("tax_report", "Tax Report", r'\bTax\s+Report\b'),
        ("prepaid_interest", "Prepaid Interest", r'\bPrepaid\s+Interest\b'),
        ("origination_fee", "Origination Fee", r'\bOrigination\s+Fee\b'),
        ("processing_fee", "Processing Fee", r'\bProcessing\s+Fee\b'),
        ("points", "Points", r'\b0\.25%\s+of\s+Loan\s+Amount\s+\(Points\)\b'),
        ("appraisal_fee", "Appraisal Fee", r'\bAppraisal\s+Fee\b'),
        ("aggregate_adjustment", "Aggregate Adjustment", r'\bAggregate\s+Adjustment\b'),
        ("homeowners_insurance_impound", "Homeowner's Insurance Impound", r"\bHomeowner'?s\s+Insurance\s+\d+\s+mo"),
        ("county_property_taxes_impound", "County Property Taxes Impound", r'\bCounty\s+Property\s+Taxes\s+\d+\s+mo'),
        ("owners_title_insurance", "Owner's Title Insurance", r"\bOwner'?s\s+Title\s+Insurance\b"),
        ("loan_policy", "Loan Policy", r'\bLoan\s+Policy\b'),
        ("recording_services", "Recording Services", r'\bRecording\s+Services\b'),
        ("escrow_fee", "Escrow Fee", r'\bEscrow\s+Fee\b'),
        ("notary_signing_fee", "Notary / Signing Fee", r'\bNotary/Signing\s+Fee\b'),
        ("special_messenger_service", "Special Messenger Service", r'\bSpecial\s+Messenger\s+Service\b'),
        ("broker_rebate", "Broker Rebate / Buyer Credit", r'\bBroker\s+Rebate\b'),
        ("county_transfer_tax", "County Documentary Transfer Tax", r'\bCounty\s+Documentary\s+Transfer\s+Tax\b'),
        ("homeowners_insurance_premium", "Homeowner's Insurance Premium", r"\bHomeowner'?s\s+Insurance\s+Premium\b"),
        ("natural_hazard_disclosure_fee", "Natural Hazard Disclosure Fee", r'\bFAN\s+HD\s+Lot\s+Report\s+Fee\b'),
    ]
    line_items = []
    if data.get("prior_loan_payoff_amount") is not None:
        line_items.append({
            "key": "prior_loan_payoff_amount",
            "label": f"Prior loan payoff to {data.get('prior_loan_payoff_lender') or 'lender'}",
            "amount": data["prior_loan_payoff_amount"],
        })
    for key, label, pattern in important_fields:
        if key in data:
            line_items.append({"key": key, "label": label, "amount": data[key]})
            continue
        if key == "loan_amount":
            value = _line_after_label_amount(text, pattern, min_value=10000, prefer_last_on_line=True)
        else:
            value = _line_after_label_amount(text, pattern)
        if value is not None:
            line_items.append({"key": key, "label": label, "amount": value})
            data.setdefault(key, value)
    if line_items:
        data["settlement_line_items"] = line_items

    return data


def parse_closing_statement(text: str) -> Dict[str, Any]:
    """Extract fields from any closing / settlement document.

    Handles three formats generically:
      - CFPB Closing Disclosure (standardised since 2015)
      - ALTA Settlement Statement (title-company form)
      - HUD-1 Settlement Statement (legacy form)

    All three share the same output schema; field names follow the app's
    Loan / Property model conventions so _apply_extracted() works unchanged.
    """
    data = parse_property_address(text)  # general address extraction first
    # Document role is derived from the document itself, not from which UI
    # button initiated the upload. Settlement-style statements establish final
    # buyer totals; a Closing Disclosure can also establish loan terms.
    if re.search(
        r'alta\s+settlement\s+statement|hud-?1\s+settlement\s+statement|'
        r'estimated\s+buyer[\'’]?s\s+statement|\bsettlement\s+statement\b',
        text,
        re.IGNORECASE,
    ):
        data['setup_import_role'] = 'settlement_document'
    else:
        data['setup_import_role'] = 'closing_document'

    purpose_match = re.search(r'\bPurpose\s*:?\s*(Purchase|Refinance|Modification)\b', text, re.IGNORECASE)
    if purpose_match:
        transaction_purpose = purpose_match.group(1).upper()
    elif re.search(r'\bPayoff\(s\)|TOTAL\s+PAYOFFS\s+AND\s+PAYMENTS', text, re.IGNORECASE):
        transaction_purpose = 'REFINANCE'
    elif re.search(r'^\s*Sale\s+Price(?:\s+of\s+Property)?\b', text, re.IGNORECASE | re.MULTILINE):
        transaction_purpose = 'PURCHASE'
    else:
        transaction_purpose = 'UNKNOWN'
    data['transaction_purpose'] = transaction_purpose
    data['loan_purpose'] = transaction_purpose.title() if transaction_purpose != 'UNKNOWN' else data.get('loan_purpose')
    data['document_type'] = 'SETTLEMENT_STATEMENT' if data['setup_import_role'] == 'settlement_document' else 'CLOSING_DISCLOSURE'
    data['transaction_role'] = (
        'ACQUISITION_SOURCE' if transaction_purpose == 'PURCHASE'
        else 'REFINANCE_SOURCE' if transaction_purpose == 'REFINANCE'
        else 'SUPPORTING_SOURCE'
    )
    data['classification_confidence'] = 0.99 if purpose_match else (0.94 if transaction_purpose != 'UNKNOWN' else 0.6)

    issued_match = re.search(r'Date\s+Issued\s*:?\s*(\d{1,2}/\d{1,2}/\d{2,4})', text, re.IGNORECASE)
    if issued_match:
        data['issued_date'] = _normalize_date(issued_match.group(1))
    disbursement_match = re.search(
        r'Disbursement\s+Date\s*:?\s*([A-Za-z]+\s+\d{1,2},?\s*\d{4}|\d{1,2}/\d{1,2}/\d{2,4})',
        text,
        re.IGNORECASE,
    )
    if disbursement_match:
        data['disbursement_date'] = _normalize_date(disbursement_match.group(1))

    # Title-company buyer/settlement statements commonly use a compact
    # "Property:" or "Property Location:" label rather than "Property Address".
    # Keep this closing-specific and line-anchored so phrases such as
    # "Sale Price of Property" cannot be mistaken for the subject address.
    m = re.search(
        r'^\s*Property(?:\s+Location)?\s*:\s*'
        r'(\d+[^\n,]*?)(?:,\s*|\n\s*)'
        r'([^,\n]+),\s*([A-Z]{2})\s+(\d{5})(?:-\d{4})?\b',
        text,
        re.IGNORECASE | re.MULTILINE,
    )
    if m:
        data['property_address'] = m.group(1).strip()
        data['property_city'] = m.group(2).strip()
        data['property_state'] = m.group(3).upper()
        data['property_zip'] = m.group(4)

    # ── Property address — CFPB "Security Interest" section is most reliable ──
    # "You are granting a security interest in\n10575 East Mission Lane, Scottsdale, AZ 85258"
    m = re.search(
        r'security\s+interest\s+in\s*\n?\s*'
        r'(\d+[^\n,]+)(?:,|\n\s*)([^,\n]+),\s*([A-Z]{2})\s+(\d{5})',
        text, re.IGNORECASE
    )
    if m:
        data['property_address'] = m.group(1).strip()
        data['property_city']    = m.group(2).strip()
        data['property_state']   = m.group(3)
        data['property_zip']     = m.group(4)

    # ── Escrow / file number ──────────────────────────────────────────────────
    m = re.search(r'file\s+(?:no\.?|#)\s*(?:/\s*escrow\s+no\.?)?\s*([A-Z0-9][-A-Z0-9/]{2,40})',
                  text, re.IGNORECASE)
    if m:
        data['escrow_number'] = m.group(1).strip()

    # ── Loan ID / reference number ────────────────────────────────────────────
    m = re.search(r'loan\s*id\s*#?\s*(\d{6,20})', text, re.IGNORECASE)
    if m:
        data['loan_id'] = m.group(1)

    # ── Borrower ─────────────────────────────────────────────────────────────
    # ALTA: "Borrower: Full Name" at start of line
    # CFPB: "... Borrower Pavani Donepudi LoanTerm 30 years" (mid-line Transaction Info)
    # Non-greedy {1,3}? prevents consuming the following "LoanTerm" keyword
    for pat in [
        r'^borrower\s*:\s*(.+)',                                             # ALTA colon
        r'\bBorrower\s+([A-Z][a-z]+(?:\s+[A-Za-z]+){1,3}?)(?=\s+(?:Loan|Purpose|Product|\d))',  # CFPB inline
        r'^Borrower\s+([A-Z][a-z]+(?:\s+[A-Za-z]+){1,3})\s*$',             # CFPB addendum line
    ]:
        m = re.search(pat, text, re.MULTILINE)
        if m:
            candidate = (m.group(1) or '').strip()
            if len(candidate) > 3 and not candidate.startswith('-'):
                data['borrower_1'] = candidate
                break

    # ── Seller ────────────────────────────────────────────────────────────────
    # ALTA: "Seller: Name" at line start   CFPB: "... Seller = Gerald I Blackman ..."
    for pat in [
        r'^seller\s*:\s*(.+)',                                               # ALTA colon
        r'\bSeller\b\s*=\s*([A-Za-z][^\n]+?)(?=\s*\n|\s{3,})',             # CFPB "= Name"
    ]:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            candidate = m.group(1).strip()
            if candidate and not candidate.lower().startswith(('credit', 'paid', 'due')):
                data['seller_name'] = candidate
                break

    # ── Lender ────────────────────────────────────────────────────────────────
    # ALTA: "Lender: Company Name" at line start
    # CFPB: "Lender Better Mortgage Corporation LoanID# ..."
    # CFPB variant: "Lender = DHi Mortgage Company, Ltd., LP. LoanID# ..."
    for pat in [
        r'^lender\s*:\s*(.+)',                                               # ALTA colon
        r'\bLender\s*=\s*([A-Za-z][^\n]+?)(?:\s+Loan(?:ID|Type|Term)|$)',   # "= Name" then LoanID or EOL
        r'\bLender\s+([A-Za-z][A-Za-z\s,.-]+?)(?=\s+Loan(?:ID|Type|Term)|\n)',  # CFPB before LoanID/etc.
    ]:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            candidate = m.group(1).strip().rstrip('.')
            # Reject single generic words like "Credits", "Charges", "Name"
            if len(candidate) > 8 and (' ' in candidate or ',' in candidate):
                data['lender_name'] = candidate
                break

    # ── Closing / Settlement date ─────────────────────────────────────────────
    # CFPB: "Closing Date 5/27/2021"   ALTA: "Settlement Date : September 20, 2024"
    date_pattern = r'(?:closing|settlement)\s+date\s*:?\s*([A-Za-z]+\s+\d{1,2},?\s*\d{4}|\d{1,2}/\d{1,2}/\d{2,4})'
    m = re.search(date_pattern, text, re.IGNORECASE)
    if m:
        normed = _normalize_date(m.group(1).strip())
        data['settlement_date']  = normed
        data['closing_date'] = normed
        if data.get('transaction_purpose') == 'PURCHASE':
            data['purchase_date'] = normed
        data['origination_date'] = normed
        if data.get('transaction_purpose') == 'PURCHASE':
            data['purchase_date_source'] = 'closing_or_settlement_date'

    if not data.get('purchase_date') and re.search(r'\bpurpose\s+purchase\b', text, re.IGNORECASE):
        m = re.search(
            r'date\s+issued\s*:?\s*([A-Za-z]+\s+\d{1,2},?\s*\d{4}|\d{1,2}/\d{1,2}/\d{2,4})',
            text,
            re.IGNORECASE,
        )
        if m:
            normed = _normalize_date(m.group(1).strip())
            data['purchase_date'] = normed
            data['origination_date'] = normed
            data['purchase_date_source'] = 'date_issued_purchase_document'

    # ── Sale / Purchase price ─────────────────────────────────────────────────
    # CFPB: "Sale Price $655,000"   ALTA: "Sale Price of Property 675,200.00"
    m = re.search(
        r'^\s*sale[^\S\n]+price(?:[^\n$]*?)?\$[^\d\n]*([\d,]+(?:\.\d{1,2})?)',
        text, re.IGNORECASE | re.MULTILINE
    )
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['purchase_price'] = v

    settlement_calculations = _extract_alta_settlement_calculations(text)
    if settlement_calculations:
        data.update({k: v for k, v in settlement_calculations.items() if v not in (None, "", [])})
        if not data.get('purchase_price') and data.get('sale_price'):
            data['purchase_price'] = data['sale_price']
        if not data.get('original_amount') and data.get('loan_amount'):
            data['original_amount'] = data['loan_amount']
        if data.get('settlement_total_amount') and data.get('purchase_price'):
            data['settlement_purchase_price_adjustment'] = round(
                float(data['settlement_total_amount']) - float(data['purchase_price']),
                2,
            )

    # ── Loan amount ───────────────────────────────────────────────────────────
    # Inline (ALTA): "Loan Amount $468,750.00" on the same line
    m = re.search(r'^\s*loan\s+amount\s+\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE | re.MULTILINE)
    if m:
        v = _parse_amount(m.group(1))
        if v and v >= 10000:
            data['original_amount'] = v
    # CFPB two-column layout: "Loan Amount" on its own line; value appears as a
    # standalone "$NNN,NNN" line several lines later (right-column flush).
    if not data.get('original_amount'):
        m = re.search(r'^loan\s+amount\s*$', text, re.IGNORECASE | re.MULTILINE)
        if m:
            rest = text[m.end():]
            val_m = re.search(r'^\$?\s*(\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?)\s*$', rest, re.MULTILINE)
            if val_m:
                v = _parse_amount(val_m.group(1))
                if v and v > 10000:
                    data['original_amount'] = v

    # ── Down payment ──────────────────────────────────────────────────────────
    # CFPB: "Down Payment/Funds from Borrower $164,000.00"
    m = re.search(r'down\s+payment[^$\n]*\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['down_payment'] = v
    # ALTA / title-company settlement statements don't print a "Down Payment"
    # line. The buyer's down payment is the earnest-money deposit plus the cash
    # brought at closing ("Buyer's funds to close"). (Apostrophe may be curly.)
    if not data.get('down_payment'):
        funds_m = re.search(
            r"buyer.?s\s+funds\s+to\s+close[^\d$\n]*\$?\s*([\d,]+(?:\.\d{1,2})?)",
            text, re.IGNORECASE,
        )
        if funds_m:
            funds_to_close = _parse_amount(funds_m.group(1))
            if funds_to_close is not None:
                data['buyer_funds_to_close'] = funds_to_close
                data.setdefault('cash_to_close', funds_to_close)
                deposit_amt = float(data.get('deposit') or 0)
                if not deposit_amt:
                    # "Deposit 40,500.00" (no colon) is parsed later; read it now.
                    dep_m = re.search(r'^deposit\s+\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE | re.MULTILINE)
                    if dep_m:
                        deposit_amt = _parse_amount(dep_m.group(1)) or 0
                data['down_payment'] = round(funds_to_close + deposit_amt, 2)
                data['down_payment_source'] = 'deposit_plus_buyer_funds_to_close'
    if (
        not data.get('original_amount')
        and data.get('purchase_price')
        and data.get('down_payment')
        and float(data.get('purchase_price') or 0) > float(data.get('down_payment') or 0)
    ):
        derived_original = round(float(data['purchase_price']) - float(data['down_payment']), 2)
        if derived_original >= 10000:
            data['original_amount'] = derived_original
            data['loan_amount'] = derived_original
            data['loan_amount_source'] = 'purchase_price_minus_down_payment'
    # ── Deposit / earnest money (ALTA) ────────────────────────────────────────
    m = re.search(r'^deposit\s+\$?\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE | re.MULTILINE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['deposit'] = v

    # ── Seller credit ─────────────────────────────────────────────────────────
    # Require a dollar sign OR comma-formatted amount to avoid matching bare line
    # numbers like "05" in "05 Seller Credit 05 Payoff of..."
    m = re.search(r'seller\s+credit\s+\$\s*([\d,]+(?:\.\d{1,2})?)', text, re.IGNORECASE)
    if m:
        v = _parse_amount(m.group(1))
        if v:
            data['seller_credit'] = v

    # ── Cash to close ─────────────────────────────────────────────────────────
    # CFPB page 3 "Summary of Transactions" has the definitive number
    cash_lines = re.findall(r'^.*cash\s+to\s+close.*$', text, re.IGNORECASE | re.MULTILINE)
    cash_values = [
        amount for line in cash_lines for amount in (
            _parse_amount(value) for value in re.findall(r'\$\s*([\d,]+(?:\.\d{1,2})?)', line)
        ) if amount is not None
    ]
    if cash_values:
        v = cash_values[-1]
        if v:
            data['cash_to_close'] = v

    total_due_matches = re.findall(
        r'(?:Total\s+)?Due\s+from\s+Borrower(?:\s+at\s+Closing)?[^\n$]*\$?\s*([\d,]+(?:\.\d{1,2})?)',
        text,
        re.IGNORECASE,
    )
    if total_due_matches:
        data['total_due_from_borrower'] = _parse_amount(total_due_matches[-1])
    total_paid_matches = re.findall(
        r'Total\s+Paid\s+Already\s+by\s+or\s+on\s+Behalf\s+of\s+Borrower[^\n$-]*-?\$?\s*([\d,]+(?:\.\d{1,2})?)',
        text,
        re.IGNORECASE,
    )
    if total_paid_matches:
        data['total_paid_on_behalf_of_borrower'] = _parse_amount(total_paid_matches[-1])
    paid_before_matches = re.findall(
        r'Closing\s+Costs\s+Paid\s+Before\s+Closing[^\n$-]*-?\$?\s*([\d,]+(?:\.\d{1,2})?)',
        text,
        re.IGNORECASE,
    )
    if paid_before_matches:
        data['deposit_paid_before_closing'] = _parse_amount(paid_before_matches[-1])

    payoff_total_lines = re.findall(
        r'^.*(?:K\.\s*)?TOTAL\s+PAYOFFS\s+AND\s+PAYMENTS.*$',
        text,
        re.IGNORECASE | re.MULTILINE,
    )
    payoff_total_values = [
        amount for line in payoff_total_lines for amount in (
            _parse_amount(value) for value in re.findall(r'\$\s*([\d,]+(?:\.\d{1,2})?)', line)
        ) if amount is not None
    ]
    if payoff_total_values:
        data['total_payoffs_and_payments'] = payoff_total_values[-1]

    # ── Borrower-paid closing costs ───────────────────────────────────────────
    # CFPB forms often show "Borrower-Paid Closing Costs" or "Closing Costs
    # Financed / Paid Before / Paid At Closing". Keep this distinct from Cash to
    # Close, which includes credits, deposits, and adjustments.
    for pat in [
        r'borrower[\s-]*paid\s+closing\s+costs?[^\n$]*\$?\s*([\d,]+(?:\.\d{1,2})?)',
        r'total\s+closing\s+costs?[^\n$]*\$?\s*([\d,]+(?:\.\d{1,2})?)',
        r'^\s*closing\s+costs?\s*\$?\s*([\d,]+(?:\.\d{1,2})?)\s*$',
        r'closing\s+costs?[^\n$]{0,60}borrower[\s-]*paid[^\n$]*\$?\s*([\d,]+(?:\.\d{1,2})?)',
    ]:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            v = _parse_amount(m.group(1))
            if v:
                data['closing_costs'] = v
                break

    # Settlement debit/credit totals are accounting controls, not closing
    # costs. Keep these semantic values independent and never derive one from
    # the other.

    # ── Interest rate ─────────────────────────────────────────────────────────
    # CFPB inline: "Interest Rate 3.625%" on same line
    m = re.search(r'interest\s+rate\s+([\d.]+)%', text, re.IGNORECASE)
    if m:
        data['interest_rate'] = float(m.group(1))
    # CFPB two-column split: "Interest Rate" on its own line, "7.625%" several lines later
    if not data.get('interest_rate'):
        m = re.search(r'^interest\s+rate\s*$', text, re.IGNORECASE | re.MULTILINE)
        if m:
            rest = text[m.end():]
            val_m = re.search(r'^([\d.]+)\s*%\s*$', rest, re.MULTILINE)
            if val_m:
                v = float(val_m.group(1))
                if 0.1 < v < 30:
                    data['interest_rate'] = v
    if not data.get('interest_rate'):
        # ALTA fallback: derive from prepaid daily interest
        m = re.search(r'prepaid\s+interest\s+\$?([\d.]+)\s+per\s+day', text, re.IGNORECASE)
        if m and data.get('original_amount'):
            daily = float(m.group(1))
            data['interest_rate'] = round(daily * 365 / data['original_amount'] * 100, 2)

    # ── Monthly P&I payment ───────────────────────────────────────────────────
    # CFPB inline: "Principal & Interest $2,239.21" on same line
    m = re.search(
        r'principal\s*[&+]\s*interest\s+\$?\s*([\d,]+\.\d{2})',
        text, re.IGNORECASE
    )
    if m:
        data['monthly_payment'] = _parse_amount(m.group(1))
    # CFPB two-column split: "Principal & Interest" on its own line, value several lines later
    if not data.get('monthly_payment'):
        m = re.search(r'^(?:monthly\s+)?principal\s*&\s*interest\s*$', text, re.IGNORECASE | re.MULTILINE)
        if m:
            rest = text[m.end():]
            val_m = re.search(r'^\$?\s*(\d{1,3}(?:,\d{3})*\.\d{2})\s*$', rest, re.MULTILINE)
            if val_m:
                v = _parse_amount(val_m.group(1))
                if v and v > 100:
                    data['monthly_payment'] = v

    # ── Loan term ─────────────────────────────────────────────────────────────
    m = re.search(r'loan\s*term\s+(\d+)\s*years?', text, re.IGNORECASE)
    if m:
        data['loan_term_years'] = int(m.group(1))

    # ── Loan type ─────────────────────────────────────────────────────────────
    if re.search(r'\bARM\b|\badjustable[\s-]rate\b', text, re.IGNORECASE):
        data['loan_type'] = 'ARM'
    elif re.search(r'fixed[\s-]rate|product\s+fixed|\bconventional\b', text, re.IGNORECASE):
        data['loan_type'] = 'FIXED'

    # ── Annual Property Tax ───────────────────────────────────────────────────
    # CFPB escrow section: "Property Taxes $682.29 per month for N mo."
    m = re.search(r'property\s+taxes?\s+\$?([\d,]+\.\d{2})\s+per\s+month', text, re.IGNORECASE)
    if m:
        data['annual_property_tax'] = round(float(m.group(1).replace(',', '')) * 12, 2)
    # ALTA impound: "Property Taxes to … N Months at $791.39/month"
    if not data.get('annual_property_tax'):
        m = re.search(r'property\s+taxes?\s+to[^\n]*?\$?([\d,]+\.\d{2})\s*/\s*month',
                      text, re.IGNORECASE)
        if m:
            data['annual_property_tax'] = round(float(m.group(1).replace(',', '')) * 12, 2)

    # ── Annual Homeowner's Insurance ──────────────────────────────────────────
    # CFPB: "Homeowner's Insurance Premium (12 mo.) to Safeco $1,113.00"
    m = re.search(
        r"homeowner.{0,3}s\s+insurance\s+premium\s*\(12\s+mo\.[^)]*\)[^\n]*?\$\s*([\d,]+\.\d{2})",
        text, re.IGNORECASE
    )
    if m:
        data['annual_insurance'] = _parse_amount(m.group(1))
    # CFPB escrow detail: "Homeowner's Insurance $92.75 per month for N mo."
    # Handle OCR artifacts like "—_" between "Insurance" and "$"
    if not data.get('annual_insurance'):
        m = re.search(r"homeowner.{0,3}s\s+insurance[^\n$]*?\$\s*([\d,]+(?:\.\d{2})?)\s+per\s+month",
                      text, re.IGNORECASE)
        if m:
            data['annual_insurance'] = round(float(m.group(1).replace(',', '')) * 12, 2)
    # ALTA: "Homeowner's Insurance Premium to … 12 months NNN.NN"
    if not data.get('annual_insurance'):
        m = re.search(r"homeowner'?s?\s+insurance\s+premium[^\n]+12\s+months?\s+([\d,]+\.\d{2})",
                      text, re.IGNORECASE)
        if m:
            data['annual_insurance'] = _parse_amount(m.group(1))
    # ALTA monthly: "Homeowner's Insurance to … $59.85/month"
    if not data.get('annual_insurance'):
        m = re.search(r"homeowner'?s?\s+insurance\s+to[^\n]*?\$?([\d,]+\.\d{2})\s*/\s*month",
                      text, re.IGNORECASE)
        if m:
            data['annual_insurance'] = round(float(m.group(1).replace(',', '')) * 12, 2)

    # ── HOA (non-escrowed recurring cost) ─────────────────────────────────────
    # CFPB: "Non-Escrowed Property Costs over Year 1 $394.92" (HOA Dues)
    m = re.search(r'non[\s-]escrowed\s+property\s+costs[^\n]*?([\d,]+\.\d{2})', text, re.IGNORECASE)
    if m:
        data['hoa_annual'] = _parse_amount(m.group(1))

    # ── Monthly escrow payment ────────────────────────────────────────────────
    # CFPB: "Estimated Escrow + 775.04"  or  "Monthly $775.04"
    m = re.search(r'estimated\s+escrow\s+\+\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
    if m:
        data['escrow_monthly'] = _parse_amount(m.group(1))

    # ── Estimated total monthly payment ───────────────────────────────────────
    m = re.search(r'estimated\s+total\s+monthly\s+payment[^\n$]*\$?\s*([\d,]+\.\d{2})', text, re.IGNORECASE)
    if m:
        data['estimated_total_monthly_payment'] = _parse_amount(m.group(1))
    elif data.get('monthly_payment') and data.get('escrow_monthly'):
        data['estimated_total_monthly_payment'] = round(data['monthly_payment'] + data['escrow_monthly'], 2)

    # ── APR (informational) ───────────────────────────────────────────────────
    m = re.search(r'annual\s+percentage\s+rate\s*\(apr\)[^\n]*?([\d.]+)%', text, re.IGNORECASE)
    if m:
        data['apr'] = float(m.group(1))

    return data


CATEGORY_TITLES = {
    'escrow_analysis': 'Annual Escrow Analysis',
    'mortgage_statement': 'Mortgage Statement',
    'tax_return': 'Tax Return',
    '1098': 'Form 1098 - Mortgage Interest Statement',
    '1099': 'Form 1099 Year-End Statement',
    'loan_disclosure': 'Loan Disclosure',
    'closing_statement': 'ALTA Settlement Statement / Closing Statement',
    'bank_statement': 'Bank Statement',
    'property_tax': 'Property Tax Statement',
    'other': 'Document',
}


def _md_table(rows: list) -> list:
    lines = ['| Field | Value |', '|---|---|']
    for label, value in rows:
        lines.append(f'| {label} | {value} |')
    lines.append('')
    return lines


def _fmt_value(key: str, v: Any) -> str:
    if isinstance(v, float):
        if 'rate' in key:
            return f'{v}%'
        return f'${v:,.2f}'
    return str(v)


def to_markdown(category: str, data: Dict[str, Any]) -> str:
    """Render extracted fields as a structured markdown document.

    Mortgage statements get the sectioned layout (Property Information,
    Payment Information, Payment Breakdown, Extraction Schema); other
    categories get a single field table.
    """
    title = CATEGORY_TITLES.get(category, 'Document')
    lines = [f'# {title} - Structured Fields', '']

    def section(source, name, fields):
        rows = [
            (label, _fmt_value(k, source[k]))
            for k, label in fields if source.get(k) is not None
        ]
        if rows:
            lines.append(f'## {name}')
            lines.append('')
            lines.extend(_md_table(rows))

    if category == 'mortgage_statement':
        display = dict(data)
        full_addr = data.get('property_address', '')
        if data.get('property_city'):
            full_addr += f", {data['property_city']}"
        if data.get('property_state'):
            full_addr += f", {data['property_state']} {data.get('property_zip', '')}".rstrip()
        if full_addr:
            display['property_address'] = full_addr
        borrowers = '; '.join(
            data[k] for k in ('borrower_1', 'borrower_2', 'borrower_3', 'borrower_4')
            if data.get(k)
        )
        if borrowers:
            display['borrowers'] = borrowers

        section(display, 'Property Information', [
            ('property_address', 'Property Address'),
            ('borrowers', 'Borrowers'),
            ('account_number', 'Mortgage Account Number'),
            ('original_amount', 'Original Principal Balance'),
            ('current_balance', 'Unpaid Principal Balance'),
            ('interest_rate', 'Interest Rate'),
            ('maturity_date', 'Maturity Date'),
        ])
        section(display, 'Payment Information', [
            ('statement_date', 'Statement Date'),
            ('payment_due_date', 'Payment Due Date'),
            ('monthly_payment', 'Amount Due'),
            ('escrow_amount', 'Escrow Amount'),
        ])
        section(display, 'Current Payment Breakdown', [
            ('principal_due', 'Principal Portion'),
            ('interest_due', 'Interest Portion'),
        ])

    if category == 'escrow_analysis':
        section(data, 'Escrow Payment', [
            ('loan_number', 'Loan Number'),
            ('property_address', 'Property Address'),
            ('statement_date', 'Statement Date'),
            ('effective_date', 'New Payment Effective Date'),
            ('current_escrow_payment', 'Current Escrow Payment'),
            ('new_escrow_payment', 'New Escrow Payment'),
        ])
        section(data, 'Annual Tax And Insurance', [
            ('history_period_start', 'History Period Start'),
            ('history_period_end', 'History Period End'),
            ('estimated_tax', 'Estimated Tax'),
            ('actual_tax', 'Actual Tax'),
            ('estimated_insurance', 'Estimated Insurance'),
            ('actual_insurance', 'Actual Insurance'),
            ('projected_tax', 'Projected Tax'),
            ('projected_insurance', 'Projected Insurance'),
            ('projected_total', 'Projected Total'),
        ])
    if category == 'closing_statement':
        display = dict(data)
        full_addr = data.get('property_address', '')
        if data.get('property_city'):
            full_addr += f", {data['property_city']}"
        if data.get('property_state'):
            full_addr += f", {data['property_state']} {data.get('property_zip', '')}".rstrip()
        if full_addr:
            display['property_address'] = full_addr

        section(display, 'Transaction Details', [
            ('escrow_number', 'Escrow / File Number'),
            ('settlement_date', 'Settlement Date'),
            ('property_address', 'Property Address'),
            ('borrower_1', 'Borrower(s)'),
            ('seller_name', 'Seller'),
            ('lender_name', 'Lender'),
        ])
        section(display, 'Purchase Financials', [
            ('purchase_price', 'Sale Price of Property'),
            ('settlement_total_amount', 'Settlement Final Total'),
            ('settlement_purchase_price_adjustment', 'Settlement Adjustment'),
            ('original_amount', 'Loan Amount'),
            ('down_payment', 'Down Payment'),
            ('deposit', 'Deposit (Earnest Money)'),
            ('seller_credit', 'Seller Credit'),
            ('cash_to_close', "Cash to Close"),
        ])
        section(display, 'Settlement Calculations', [
            ('settlement_debit_subtotal', 'Buyer Debit Subtotal'),
            ('settlement_credit_subtotal', 'Buyer Credit Subtotal'),
            ('settlement_due_to_buyer', 'Due To Buyer'),
            ('settlement_debit_total', 'Buyer Debit Total'),
            ('settlement_credit_total', 'Buyer Credit Total'),
            ('settlement_total_amount', 'Final Settlement Total'),
        ])
        section(display, 'Loan Terms', [
            ('interest_rate', 'Interest Rate'),
            ('monthly_payment', 'Monthly Principal & Interest Payment'),
            ('loan_term_years', 'Loan Term (years)'),
            ('loan_type', 'Loan Type'),
            ('apr', 'APR'),
        ])
        section(display, 'Operating Cost Estimates', [
            ('annual_property_tax', 'Annual Property Tax'),
            ('annual_insurance', 'Annual Homeowner\'s Insurance'),
            ('escrow_monthly', 'Monthly Escrow (taxes + insurance)'),
            ('hoa_annual', 'Annual HOA Dues (non-escrowed)'),
        ])

    if category == '1098':
        display = dict(data)
        full_addr = data.get('property_address', '')
        if data.get('property_city'):
            full_addr += f", {data['property_city']}"
        if data.get('property_state'):
            full_addr += f", {data['property_state']} {data.get('property_zip', '')}".rstrip()
        if full_addr:
            display['property_address'] = full_addr
        borrowers = '; '.join(
            data[k] for k in ('borrower_1', 'borrower_2', 'borrower_3', 'borrower_4')
            if data.get(k)
        )
        if borrowers:
            display['borrowers'] = borrowers

        section(display, 'Statement Information', [
            ('lender_name', 'Lender / Recipient'),
            ('borrowers', 'Borrowers'),
            ('account_number', 'Account Number'),
            ('tax_year', 'Calendar Year'),
            ('property_address', 'Property Securing Mortgage'),
        ])
        section(display, 'Mortgage Interest (Form 1098)', [
        ('mortgage_interest', 'Box 1 - Mortgage Interest Received'),
        ('current_balance', 'Box 2 - Outstanding Principal'),
        ('origination_date', 'Box 3 - Origination Date'),
        ('mortgage_insurance', 'Box 5 - Mortgage Insurance Premiums'),
        ('points_paid', 'Box 6 - Points Paid'),
        ('property_tax_amount', 'Box 10 - Real Estate Taxes Paid'),
        ('mortgage_acquisition_date', 'Box 11 - Mortgage Acquisition Date'),
    ])

    # Raw key/value schema — what the app imports
    schema_rows = [
        (k, f'{v:,.2f}' if isinstance(v, float) and 'rate' not in k else v)
        for k, v in data.items()
        if k != 'raw_text_preview' and v is not None
    ]
    if schema_rows:
        lines.append('## Recommended Extraction Schema for App Import')
        lines.append('')
        lines.extend(_md_table(schema_rows))

    return '\n'.join(lines)


def parse_document(filepath: str, category: str = 'auto') -> tuple[str, Dict[str, Any], str]:
    """Main entry: parse document, returning (category, extracted data, markdown).

    When category is "auto", PDF documents are detected from extracted text.
    """
    path = Path(filepath)
    ext = path.suffix.lower()

    raw_data: Dict[str, Any] = {}

    if ext == '.pdf':
        text = extract_pdf_text(filepath)
        if not text or not text.strip():
            raise ValueError("Document text extraction produced no readable text.")
        if category == 'auto':
            category = detect_category(text)

        if category == 'escrow_analysis':
            raw_data = parse_escrow_analysis(text)
        elif category == 'mortgage_statement':
            raw_data = parse_mortgage_statement(text)
            _log_loan_document_extraction_gaps(category, raw_data, text, filepath)
        elif category == 'tax_return':
            # Per-property Schedule E figures are returned in the properties
            # collection so the preview can show field-level mappings.
            tr = parse_tax_return_properties(filepath)
            rentals = [
                p for p in tr.get('properties', [])
                if p.get('property_kind') == 'rental'
            ]
            raw_data = {
                'tax_year': tr.get('tax_year'),
                'statement_year': tr.get('tax_year'),
                'property_count': len(tr.get('properties', [])),
                'rental_count': len(rentals),
                'total_rents_received': round(
                    sum(p.get('rents_received') or 0 for p in rentals), 2
                ),
                'total_mortgage_interest': round(
                    sum(p.get('mortgage_interest') or 0 for p in rentals), 2
                ),
                'schedule1_line5_total': tr.get('schedule1_line5_total'),
                'schedule1_line5_delta': tr.get('schedule1_line5_delta'),
                'schedule1_line5_warning': tr.get('schedule1_line5_warning'),
                'form4562_present': tr.get('form4562_present'),
                'depreciation_worksheet_present': tr.get('depreciation_worksheet_present'),
                'properties': tr.get('properties', []),
                'period_type': 'yearly',
            }
        elif category == '1098':
            # The IRS Form 1098 box grid often lives on a later page of a
            # year-end packet that MarkItDown drops. Detect the page first: if
            # the primary text has no Box 1 label, use pdfplumber's positional
            # text, which preserves every page.
            form_text = text
            if not _FORM_1098_START.search(text):
                try:
                    alt = _pdfplumber_text(filepath, x_tolerance=1)
                    if _FORM_1098_START.search(alt):
                        form_text = alt
                except Exception:
                    pass
            raw_data = parse_1098(form_text)
            _log_loan_document_extraction_gaps(category, raw_data, text, filepath)
        elif category == '1099':
            raw_data = parse_1099(text)
        elif category == 'closing_statement':
            # Purchase Closing Disclosures serve two workflows: Property Setup
            # uses sale/closing values, while Loan Setup uses their debt terms.
            # Keep the category stable and enrich the extracted payload.
            raw_data = (
                parse_loan_disclosure(text)
                if re.search(r'closing\s+disclosure', text, re.IGNORECASE)
                else parse_closing_statement(text)
            )
        elif category == 'loan_disclosure':
            raw_data = parse_loan_disclosure(text)
        elif category == 'bank_statement':
            raw_data = parse_bank_statement(text)
        elif category == 'property_tax':
            raw_data = parse_property_tax(text)
        elif category == 'insurance_declaration':
            raw_data = parse_insurance_declaration(text)
        else:
            raw_data = {'raw_text_preview': text[:500]}

        if 'raw_text_preview' not in raw_data:
            raw_data['period_type'] = raw_data.get('period_type') or detect_period_type(raw_data)
    elif ext in ('.xlsx', '.xls'):
        text = _markitdown_text(filepath)
        if text and len(text.strip()) > 30:
            detected = detect_category(text) if category == 'auto' else category
            if detected == 'mortgage_statement':
                category = detected
                raw_data = parse_mortgage_statement(text)
                _log_loan_document_extraction_gaps(category, raw_data, text, filepath)
            elif detected == '1098':
                category = detected
                raw_data = parse_1098(text)
                _log_loan_document_extraction_gaps(category, raw_data, text, filepath)
            else:
                category = 'other' if category == 'auto' else category
                raw_data = extract_excel_data(filepath)
        else:
            logger.warning("MarkItDown produced empty text for spreadsheet upload %s", filepath)
            if category == 'auto':
                category = 'other'
            raw_data = extract_excel_data(filepath)
    elif category == 'auto':
        category = 'other'

    markdown = to_markdown(category, raw_data) if isinstance(raw_data, dict) else ''
    return category, raw_data, markdown
